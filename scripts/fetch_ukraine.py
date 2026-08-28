#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""fetch_ukraine.py — vue « Guerre en Ukraine » de l'Atlas Économique.

Règle de la vue : ne publier que des grandeurs MESURÉES ou PUBLIÉES par une
institution qui les signe, jamais une estimation maison. Chaque nombre du cache
porte sa source et sa date. Ce que personne ne publie proprement (pertes
humaines militaires, PIB russe réel, budget de guerre russe) n'entre PAS ici.

Six blocs, six sources indépendantes :

  1. FRONT — DeepStateMap (deepstatemap.live/api). On ne recopie pas un chiffre
     de territoire occupé : on télécharge le GeoJSON du front et on CALCULE
     l'aire nous-mêmes (excès sphérique). Contrôle de vraisemblance intégré :
     la Crimée doit retomber sur ~26 900 km² (chiffre officiel ukrainien) ;
     si l'écart dépasse 5 %, le bloc est refusé. La série historique est
     reconstituée snapshot par snapshot et conservée d'un run à l'autre.

  2. AIDE — Ukraine Support Tracker de l'IfW Kiel (classeur XLSX). Le lien du
     classeur change à chaque « Release » : on le RETROUVE sur la page, on ne
     le code pas en dur. Les montants sont ceux de Kiel, en Md€, et le
     classement en % du PIB du donateur — le seul honnête — vient de la même
     feuille.

  3. NBU — Banque nationale d'Ukraine : hryvnia, réserves de change mensuelles.

  4. RÉFUGIÉS — HCR (api.unhcr.org), réfugiés d'origine ukrainienne par année.

  5. DÉCOUPLAGE ÉNERGÉTIQUE — Eurostat : importations UE de gaz et de pétrole
     en provenance de Russie, mensuel. Donnée européenne, pas russe.

  6. COMMERCE — Eurostat : échanges de marchandises UE↔Ukraine, mensuel.
     (Les escales portuaires d'IMF PortWatch ont été essayées puis ÉCARTÉES :
      leur série ukrainienne attribue le trafic aux mauvais ports depuis 2025.
      Le motif chiffré est conservé en commentaire au-dessus de build_trade.)

Produit `ukraine_cache.{json,js}` (window.__ATLAS_UKRAINE__), chargé en lazy
par la vue. Le PIB, l'inflation, la dette de l'Ukraine et de la Russie ne sont
PAS recollectés : la vue les lit dans atlas_econ_cache (FMI WEO), déjà chargé.

MODES :
  --full   tout, y compris le rattrapage de l'historique du front
  --live   (défaut) : front = derniers snapshots manquants uniquement, le reste
           rafraîchi si le cache a plus de MAXAGE_H heures
  --front  seulement le bloc front (debug)
"""
import os
import sys
import re
import json
import time
import math
import zipfile
import datetime as dt
from urllib.parse import urljoin

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from curl_cffi import requests as rq

    def _sess():
        return rq.Session(impersonate="chrome120", timeout=60)
except Exception:  # pragma: no cover - repli
    import requests as rq

    def _sess():
        s = rq.Session()
        s.headers.update({"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) "
                                        "AppleWebKit/537.36 Chrome/126 Safari/537.36"})
        return s

HOME = os.path.expanduser("~")
REPO = os.path.join(HOME, "Desktop", "Site_Crypto_Finance")
CACHES = [
    os.path.join(REPO, "ukraine_cache"),
    os.path.join(REPO, "public", "ukraine_cache"),
    os.path.join(HOME, "Library", "Caches", "site_crypto_finance", "ukraine_cache"),
]

# Superficie de l'Ukraine dans ses frontières de 1991 (dont Crimée et mer d'Azov
# intérieure exclue) — chiffre officiel du Service d'État du cadastre, repris par
# la Banque mondiale (AG.SRF.TOTL.K2). Sert de dénominateur au « % du territoire ».
UKRAINE_KM2 = 603_550.0
# Contrôle de vraisemblance du calcul d'aire : République autonome de Crimée
# (26 081 km²) + Sébastopol (864 km²) = 26 945 km².
CRIMEA_REF_KM2 = 26_945.0
CRIMEA_TOL = 0.05

MAXAGE_H = 20.0         # au-delà, --live rafraîchit aussi les blocs lents (Kiel = 12 Mo)
MAX_NEW_SNAPSHOTS = 60  # plafond de snapshots DeepStateMap téléchargés par run

DSM_API = "https://deepstatemap.live/api"
KIEL_PAGE = "https://www.ifw-kiel.de/topics/war-against-ukraine/ukraine-support-tracker/"
EUROSTAT = "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/"

WARN = []


def warn(msg):
    WARN.append(msg)
    sys.stderr.write("[WARN] %s\n" % msg)


def log(msg):
    print(msg, flush=True)


# ═══════════════════════════════════════════════════════════════════════════
# Géométrie : aire d'un polygone sur la sphère (excès sphérique).
#
# Formule classique (identique à celle de la Google Maps Geometry Library) :
#   A = R² / 2 · Σ (λ₂ − λ₁)(2 + sin φ₁ + sin φ₂)
# Elle est exacte pour un polygone sphérique et ne demande AUCUNE dépendance
# (ni shapely ni pyproj, absents de l'environnement de collecte). L'erreur
# vis-à-vis de l'ellipsoïde WGS84 est de l'ordre de 0,2 % à ces latitudes —
# négligeable devant la précision du tracé du front lui-même.
# ═══════════════════════════════════════════════════════════════════════════
R_EARTH_KM = 6371.0088


def ring_km2(ring):
    n = len(ring)
    if n < 3:
        return 0.0
    tot = 0.0
    for i in range(n):
        lo1 = math.radians(ring[i][0]); la1 = math.radians(ring[i][1])
        j = (i + 1) % n
        lo2 = math.radians(ring[j][0]); la2 = math.radians(ring[j][1])
        tot += (lo2 - lo1) * (2.0 + math.sin(la1) + math.sin(la2))
    return abs(tot * R_EARTH_KM * R_EARTH_KM / 2.0)


def geom_km2(g):
    t = (g or {}).get("type")
    c = (g or {}).get("coordinates")
    if t == "Polygon" and c:
        return ring_km2(c[0]) - sum(ring_km2(r) for r in c[1:])
    if t == "MultiPolygon" and c:
        s = 0.0
        for p in c:
            if p:
                s += ring_km2(p[0]) - sum(ring_km2(r) for r in p[1:])
        return s
    return 0.0


# ── Classement des polygones DeepStateMap ────────────────────────────────────
# TROIS FORMATS DE NOMMAGE COHABITENT dans l'historique, et il faut les lire tous
# les trois — sinon les relevés anciens sortent à zéro et quittent la série sans
# bruit (mesuré : 24 relevés sur 103 évaporés au premier essai) :
#
#   2022        « Окуповано »                          ukrainien seul
#   2022-2023   « Окуповано /// Occupied »             + anglais
#   depuis 2024 « … /// geoJSON.status.occupied »      + clé stable
#
# La clé fait foi quand elle est là. Sinon on lit le texte, dans un ORDRE qui
# compte : « Окупований Крим » contient « окупован », donc la Crimée doit être
# reconnue AVANT l'occupation générale, sans quoi elle serait comptée deux fois.
#
# ── ET LE FICHIER NE PARLE PAS QUE D'UKRAINE ─────────────────────────────────
# DeepStateMap y publie aussi des calques militants sur d'autres territoires
# qu'il considère occupés par la Russie : Kouriles, Carélie, Itchkérie, Prusse
# orientale, Abkhazie, Petsamo, Salla, district de Tskhinvali, Transnistrie,
# île de Touzla, Pechorsky. Presque tous portent le mot « occupé » : lus
# naïvement, ils ajoutent 130 000 km² au territoire ukrainien occupé.
#
# Ils sont écartés par DEUX filets indépendants, parce qu'aucun des deux ne
# suffit seul : une CLÔTURE GÉOGRAPHIQUE (le polygone doit être en Ukraine)
# élimine tout ce qui est loin, mais pas la Transnistrie ni Touzla, qui sont
# à la porte ; une liste de noms écartés s'occupe de ces derniers, mais elle
# périmerait seule au premier calque nouveau. Les deux ensemble tiennent.
KEY_OCCUPIED = "status.occupied"          # occupé depuis février 2022
KEY_CRIMEA = "territories.crimea"         # occupé depuis 2014
KEY_ORDLO = "territories.ordlo"           # Donbass séparatiste d'avant 2022
KEY_LIBERATED = "status.dismissed"        # repris par l'Ukraine
KEY_UNKNOWN = "status.unknown"            # zone grise revendiquée des deux côtés

CONTROLLED = (KEY_OCCUPIED, KEY_CRIMEA, KEY_ORDLO)

# Clôture : boîte englobante généreuse de l'Ukraine de 1991 (Crimée comprise).
UA_BBOX = (21.5, 43.3, 41.5, 53.0)        # lon min, lat min, lon max, lat max

# Calques hors sujet qui tombent DANS la clôture (voisinage immédiat).
NOT_UKRAINE = (
    "придністров", "transnistria", "тузла", "tuzla", "молдов", "moldova",
)

# Reconnaissance par le texte, dans l'ordre. Le premier motif qui apparaît gagne.
NAME_RULES = (
    (KEY_CRIMEA, ("крим", "crimea")),
    (KEY_ORDLO, ("ордло", "cadr", "calr", "ордил")),
    (KEY_LIBERATED, ("звільнено", "liberated", "звiльнено")),
    (KEY_UNKNOWN, ("невідом", "невiдом", "під питанням", "пiд питанням", "unknown",
                   "situation is unknown", "статус невідомий")),
    (KEY_OCCUPIED, ("окупован", "окуповано", "occupied")),
)


def geom_center(g):
    """Centre de la boîte englobante d'une géométrie. Suffisant pour une
    clôture : on cherche à savoir si l'objet est en Ukraine ou en Sibérie, pas
    à le localiser au mètre."""
    xs, ys = [], []

    def walk(c):
        if isinstance(c, (list, tuple)):
            if c and isinstance(c[0], (int, float)):
                xs.append(float(c[0])); ys.append(float(c[1]))
            else:
                for x in c:
                    walk(x)
    walk((g or {}).get("coordinates"))
    if not xs:
        return None
    return ((min(xs) + max(xs)) / 2.0, (min(ys) + max(ys)) / 2.0)


def in_ukraine(g):
    c = geom_center(g)
    if not c:
        return False
    return (UA_BBOX[0] <= c[0] <= UA_BBOX[2]) and (UA_BBOX[1] <= c[1] <= UA_BBOX[3])


def dsm_key(feat):
    """Catégorie d'un polygone DeepStateMap, ou None s'il ne nous concerne pas."""
    props = (feat or {}).get("properties") or {}
    nm = " ".join(str(props.get("name") or "").split())
    low = nm.lower()
    if any(w in low for w in NOT_UKRAINE):
        return None
    if not in_ukraine((feat or {}).get("geometry")):
        return None
    if "geoJSON." in nm:
        k = nm.split("geoJSON.", 1)[1].strip().split()[0].split("{{")[0].strip()
        if k.startswith("status.dismissed_at"):
            return KEY_LIBERATED
        return k or None
    for key, pats in NAME_RULES:
        if any(pt in low for pt in pats):
            return key
    return None


def dsm_areas(fc):
    """{catégorie: km²} pour une FeatureCollection DeepStateMap."""
    out = {}
    for f in (fc or {}).get("features") or []:
        g = f.get("geometry") or {}
        if g.get("type") not in ("Polygon", "MultiPolygon"):
            continue
        k = dsm_key(f)
        if not k:
            continue
        out[k] = out.get(k, 0.0) + geom_km2(g)
    return out


def dsm_totals(areas):
    occ = areas.get(KEY_OCCUPIED, 0.0)
    cri = areas.get(KEY_CRIMEA, 0.0)
    ord_ = areas.get(KEY_ORDLO, 0.0)
    return {
        "occ": round(occ, 1),
        "cri": round(cri, 1),
        "ordlo": round(ord_, 1),
        "lib": round(areas.get(KEY_LIBERATED, 0.0), 1),
        "unk": round(areas.get(KEY_UNKNOWN, 0.0), 1),
        "tot": round(occ + cri + ord_, 1),
    }


# ── QUAND UN RELEVÉ EST-IL LISIBLE ? ─────────────────────────────────────────
# Les exports anciens de l'API sont PARTIELS : certains ne portent qu'une partie
# des calques. Relevé sur des sondages trimestriels :
#
#     2022-04-03   occupé  57 886   Crimée      0   ORDLO 16 935
#     2022-07-01   occupé  16 637   Crimée      0   ORDLO      0     ← inexploitable
#     2022-10-01   occupé  71 414   Crimée 27 111   ORDLO 16 760
#     2023-01-01   occupé  64 595   Crimée 27 111   ORDLO 16 756     (retrait de Kherson)
#     …           puis une série continue et cohérente jusqu'à aujourd'hui
#
# Le critère d'acceptation ne peut donc PAS être un seuil choisi à la main : il
# doit sortir de la donnée. C'est la CRIMÉE qui le fournit. Elle est occupée sans
# interruption depuis 2014 et sa surface ne bouge pas d'un mètre ; un fichier qui
# ne la porte pas n'énonce pas un fait, il lui manque une couche. Un relevé sans
# calque Crimée — ou dont la Crimée ne retombe pas sur ses 26 945 km² — est donc
# REFUSÉ, et le motif est écrit. Ce seul critère fait commencer la série en
# octobre 2022, exactement là où elle devient cohérente : personne n'a eu à
# choisir cette date.
#
# Les bornes d'occupation ci-dessous ne sont qu'un second filet, très large.
OCC_MIN, OCC_MAX = 30_000.0, 250_000.0


def snapshot_ok(t):
    if not t["cri"]:
        return "export partiel : le calque Crimée est absent du fichier"
    if abs(t["cri"] - CRIMEA_REF_KM2) / CRIMEA_REF_KM2 > CRIMEA_TOL:
        return "Crimée mesurée à %.0f km² au lieu de %.0f" % (t["cri"], CRIMEA_REF_KM2)
    if not t["ordlo"]:
        return "export partiel : le calque du Donbass séparatiste est absent"
    if not (OCC_MIN <= t["occ"] <= OCC_MAX):
        return "occupation mesurée à %.0f km², hors des bornes plausibles" % t["occ"]
    return None


# ═══════════════════════════════════════════════════════════════════════════
# 1 · FRONT — DeepStateMap
# ═══════════════════════════════════════════════════════════════════════════
def month_anchors(hist):
    """Choisit les snapshots à mesurer : un par mois depuis le début, puis un
    par semaine sur les 14 derniers mois (le front bouge trop pour du mensuel),
    plus le tout dernier. Renvoie [(id, date_iso)] trié."""
    rows = []
    for h in hist:
        ts = h.get("createdAt") or h.get("updatedAt")
        if not ts:
            continue
        try:
            d = dt.datetime.strptime(ts[:10], "%Y-%m-%d").date()
        except Exception:
            continue
        rows.append((d, str(h.get("id"))))
    rows.sort()
    if not rows:
        return []
    last_d = rows[-1][0]
    weekly_from = last_d - dt.timedelta(days=430)
    picked = {}
    for d, i in rows:
        bucket = ("W%s-%s" % d.isocalendar()[:2]) if d >= weekly_from else d.strftime("M%Y-%m")
        picked.setdefault(bucket, (d, i))   # le PREMIER snapshot du bucket
    sel = sorted(picked.values())
    if sel[-1][1] != rows[-1][1]:
        sel.append(rows[-1])
    return [(i, d.isoformat()) for d, i in sel]


def build_front(sess, prev, full=False):
    r = sess.get(DSM_API + "/history/public")
    r.raise_for_status()
    hist = r.json()
    if not isinstance(hist, list) or not hist:
        raise RuntimeError("historique DeepStateMap vide")
    anchors = month_anchors(hist)
    log("front · %d snapshots publiés, %d points de mesure retenus" % (len(hist), len(anchors)))

    known = {}
    for row in ((prev or {}).get("front") or {}).get("series") or []:
        known[str(row.get("id"))] = row
    # Un export partiel le restera : le fichier est figé côté source. Sans
    # mémoire des refus, chaque passage re-télécharge les mêmes treize relevés
    # illisibles — pour rien, et en masquant le vrai rattrapage sous le plafond.
    skipped = set(((prev or {}).get("front") or {}).get("skipped") or [])
    todo = [(i, d) for i, d in anchors if i not in known and i not in skipped]
    if not full and len(todo) > MAX_NEW_SNAPSHOTS:
        # On rattrape par la FIN (les points récents d'abord) : une vue à trous
        # anciens reste lisible, une vue sans le présent ne l'est pas.
        todo = todo[-MAX_NEW_SNAPSHOTS:]
        warn("front · rattrapage plafonné à %d snapshots ce run (relancer pour compléter)"
             % MAX_NEW_SNAPSHOTS)

    series = list(known.values())
    last_fc = None
    last_id = anchors[-1][0]
    for n, (sid, day) in enumerate(todo, 1):
        try:
            rr = sess.get("%s/history/%s/geojson" % (DSM_API, sid))
            rr.raise_for_status()
            fc = rr.json()
            a = dsm_areas(fc)
            t = dsm_totals(a)
            bad = snapshot_ok(t)
            if bad:
                warn("front · relevé %s (%s) refusé : %s" % (day, sid, bad))
                skipped.add(sid)
                continue
            t["id"] = sid
            t["d"] = day
            series.append(t)
            if sid == last_id:
                last_fc = fc
            if n % 10 == 0 or n == len(todo):
                log("  … %d/%d (%s : %s km² sous contrôle russe)"
                    % (n, len(todo), day, format(int(t["tot"]), ",d").replace(",", " ")))
            time.sleep(0.35)
        except Exception as e:
            warn("front · snapshot %s (%s) : %s" % (sid, day, e))

    if last_fc is None:
        rr = sess.get(DSM_API + "/history/last")
        rr.raise_for_status()
        j = rr.json()
        last_fc = j.get("map") or j
        last_id = str(j.get("id") or last_id)

    # ── Contrôle de vraisemblance : la Crimée est une constante géographique.
    # Si notre calcul d'aire s'en écarte, c'est le calcul (ou le format amont)
    # qui a bougé, pas le monde : on refuse le bloc plutôt que de publier faux.
    ref = dsm_totals(dsm_areas(last_fc))
    if ref["cri"] <= 0:
        raise RuntimeError("contrôle Crimée : polygone absent du dernier snapshot")
    err = abs(ref["cri"] - CRIMEA_REF_KM2) / CRIMEA_REF_KM2
    if err > CRIMEA_TOL:
        raise RuntimeError("contrôle Crimée : %.0f km² calculés vs %.0f km² officiels "
                           "(%.1f %% d'écart) — calcul d'aire refusé"
                           % (ref["cri"], CRIMEA_REF_KM2, err * 100))
    log("front · contrôle Crimée OK : %.0f km² calculés vs %.0f officiels (%.2f %%)"
        % (ref["cri"], CRIMEA_REF_KM2, err * 100))

    series = [s for s in series if s.get("d") and s.get("occ")]
    series.sort(key=lambda s: s["d"])
    # dédoublonnage par jour (on garde la dernière mesure du jour)
    ded = {}
    for s in series:
        ded[s["d"]] = s
    series = [ded[k] for k in sorted(ded)]

    last = dict(series[-1]) if series else dict(ref)
    last["id"] = last_id
    last["pct"] = round(100.0 * last["tot"] / UKRAINE_KM2, 2)

    def delta(days):
        target = dt.date.fromisoformat(last["d"]) - dt.timedelta(days=days)
        cand = [s for s in series if dt.date.fromisoformat(s["d"]) <= target]
        if not cand:
            return None
        base = cand[-1]
        return {"from": base["d"], "km2": round(last["tot"] - base["tot"], 1)}

    hist_last = hist[-1] if hist else {}
    return {
        "last": last,
        "series": series,
        "deltas": {"d30": delta(30), "d90": delta(90), "d365": delta(365)},
        "skipped": sorted(skipped),
        "note_en": (hist_last.get("descriptionEn") or "")[:400],
        "ukraine_km2": UKRAINE_KM2,
        "crimea_ref_km2": CRIMEA_REF_KM2,
        "crimea_calc_km2": ref["cri"],
    }


# ═══════════════════════════════════════════════════════════════════════════
# 2 · AIDE À L'UKRAINE — Ukraine Support Tracker (IfW Kiel)
# ═══════════════════════════════════════════════════════════════════════════
FR_COUNTRY = {
    "Australia": "Australie", "Austria": "Autriche", "Belgium": "Belgique",
    "Bulgaria": "Bulgarie", "Canada": "Canada", "Croatia": "Croatie",
    "Cyprus": "Chypre", "Czech Republic": "Tchéquie", "Czechia": "Tchéquie",
    "Denmark": "Danemark", "Estonia": "Estonie", "Finland": "Finlande",
    "France": "France", "Germany": "Allemagne", "Greece": "Grèce",
    "Hungary": "Hongrie", "Iceland": "Islande", "India": "Inde",
    "Ireland": "Irlande", "Italy": "Italie", "Japan": "Japon",
    "Latvia": "Lettonie", "Lithuania": "Lituanie", "Luxembourg": "Luxembourg",
    "Malta": "Malte", "Netherlands": "Pays-Bas", "New Zealand": "Nouvelle-Zélande",
    "Norway": "Norvège", "Poland": "Pologne", "Portugal": "Portugal",
    "Romania": "Roumanie", "Slovakia": "Slovaquie", "Slovenia": "Slovénie",
    "South Korea": "Corée du Sud", "Korea, South": "Corée du Sud",
    "Spain": "Espagne", "Sweden": "Suède", "Switzerland": "Suisse",
    "Taiwan": "Taïwan", "Turkey": "Turquie", "Türkiye": "Turquie",
    "United Kingdom": "Royaume-Uni", "United States": "États-Unis",
    "China": "Chine", "EU (Commission and Council)": "UE (Commission et Conseil)",
    "EU Institutions": "Institutions de l'UE",
    "European Investment Bank": "Banque européenne d'investissement",
    "Israel": "Israël", "Chile": "Chili", "Albania": "Albanie",
    "North Macedonia": "Macédoine du Nord", "Montenegro": "Monténégro",
    "Serbia": "Serbie", "Moldova": "Moldavie", "Georgia": "Géorgie",
    "Bosnia and Herzegovina": "Bosnie-Herzégovine", "Kosovo": "Kosovo",
    "Liechtenstein": "Liechtenstein", "Monaco": "Monaco", "Andorra": "Andorre",
    "San Marino": "Saint-Marin",
}

ISO3 = {
    "Australie": "AUS", "Autriche": "AUT", "Belgique": "BEL", "Bulgarie": "BGR",
    "Canada": "CAN", "Croatie": "HRV", "Chypre": "CYP", "Tchéquie": "CZE",
    "Danemark": "DNK", "Estonie": "EST", "Finlande": "FIN", "France": "FRA",
    "Allemagne": "DEU", "Grèce": "GRC", "Hongrie": "HUN", "Islande": "ISL",
    "Inde": "IND", "Irlande": "IRL", "Italie": "ITA", "Japon": "JPN",
    "Lettonie": "LVA", "Lituanie": "LTU", "Luxembourg": "LUX", "Malte": "MLT",
    "Pays-Bas": "NLD", "Nouvelle-Zélande": "NZL", "Norvège": "NOR",
    "Pologne": "POL", "Portugal": "PRT", "Roumanie": "ROU", "Slovaquie": "SVK",
    "Slovénie": "SVN", "Corée du Sud": "KOR", "Espagne": "ESP", "Suède": "SWE",
    "Suisse": "CHE", "Taïwan": "TWN", "Turquie": "TUR", "Royaume-Uni": "GBR",
    "États-Unis": "USA", "Chine": "CHN", "Israël": "ISR", "Chili": "CHL",
    "Albanie": "ALB", "Macédoine du Nord": "MKD", "Monténégro": "MNE",
    "Serbie": "SRB", "Moldavie": "MDA", "Géorgie": "GEO",
    "Bosnie-Herzégovine": "BIH", "Kosovo": "XKX", "Liechtenstein": "LIE",
}


def kiel_xlsx_url(sess):
    """Le classeur porte un UUID et un numéro de release qui changent à chaque
    publication : on lit le lien sur la page plutôt que de le figer."""
    r = sess.get(KIEL_PAGE)
    r.raise_for_status()
    html = r.text
    cands = re.findall(r'href="([^"]+\.xlsx)"', html, re.I)
    best = None
    for h in cands:
        if "ukraine" in h.lower() and "support" in h.lower() and "tracker" in h.lower():
            best = h
            break
    if not best and cands:
        best = cands[0]
    if not best:
        raise RuntimeError("aucun .xlsx trouvé sur la page du tracker Kiel")
    url = urljoin(KIEL_PAGE, best)
    m = re.search(r"Release[_ ]?(\d+)", best, re.I)
    return url, (m.group(1) if m else None)


def _cell(v):
    try:
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else f
    except Exception:
        return None


def build_aid(sess):
    import openpyxl  # présent dans requirements.txt du dépôt de collecte

    url, release = kiel_xlsx_url(sess)
    log("aide · classeur Kiel release %s" % (release or "?"))
    r = sess.get(url)
    r.raise_for_status()
    raw = r.content
    if len(raw) < 200_000:
        raise RuntimeError("classeur Kiel tronqué (%d octets)" % len(raw))
    tmp = "/tmp/_kiel_ust.xlsx"
    with open(tmp, "wb") as f:
        f.write(raw)
    if not zipfile.is_zipfile(tmp):
        raise RuntimeError("classeur Kiel illisible (téléchargement corrompu)")

    wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
    ws = wb["Country Summary (€)"]
    rows = list(ws.iter_rows(values_only=True))

    hdr_i = None
    for i, row in enumerate(rows[:30]):
        if row and str(row[1] or "").strip() == "Country":
            hdr_i = i
            break
    if hdr_i is None:
        raise RuntimeError("en-tête « Country » introuvable dans la feuille Kiel")

    # Colonnes repérées PAR LEUR TITRE (l'ordre peut changer d'une release
    # à l'autre) : on cherche le couple (libellé, unité) sur les deux lignes.
    head = [str(c or "").strip() for c in rows[hdr_i]]
    unit = [str(c or "").strip() for c in rows[hdr_i + 1]]

    def col(label, u):
        for i, (h, uu) in enumerate(zip(head, unit)):
            if h == label and uu == u:
                return i
        return None

    C = {
        "fin": col("Financial allocations", "€ billion"),
        "hum": col("Humanitarian allocations", "€ billion"),
        "mil": col("Military allocations", "€ billion"),
        "tot": col("Total bilateral allocations", "€ billion"),
        "com": col("Total bilateral commitments", "€ billion"),
        "pct": col("Total bilateral allocations", "% 2021 GDP"),
        "pct_com": col("Total bilateral commitments", "% 2021 GDP"),
        "eu_tot": col("Total bilateral and EU allocations", "€ billion"),
        "eu_pct": col("Total bilateral and EU allocations", "% 2021 GDP"),
        "eumem": None, "geoeu": None,
    }
    for i, h in enumerate(head):
        if h == "EU member":
            C["eumem"] = i
        elif h == "Geographic Europe":
            C["geoeu"] = i
    missing = [k for k, v in C.items() if v is None and k not in ("eumem", "geoeu")]
    if missing:
        raise RuntimeError("colonnes Kiel introuvables : %s" % ", ".join(missing))

    out = []
    for row in rows[hdr_i + 2:]:
        if not row or not row[1]:
            continue
        name = str(row[1]).strip()
        if not name or name.lower().startswith("total"):
            continue
        tot = _cell(row[C["tot"]])
        if tot is None:
            continue
        fr = FR_COUNTRY.get(name, name)
        out.append({
            "n": name, "fr": fr, "iso3": ISO3.get(fr),
            "eu": 1 if _cell(row[C["eumem"]]) else 0,
            "geoeu": 1 if (C["geoeu"] is not None and _cell(row[C["geoeu"]])) else 0,
            "mil": round(_cell(row[C["mil"]]) or 0.0, 3),
            "fin": round(_cell(row[C["fin"]]) or 0.0, 3),
            "hum": round(_cell(row[C["hum"]]) or 0.0, 3),
            "tot": round(tot, 3),
            "com": round(_cell(row[C["com"]]) or 0.0, 3),
            "pct": round((_cell(row[C["pct"]]) or 0.0), 4),
            "pctc": round((_cell(row[C["pct_com"]]) or 0.0), 4),
            "eutot": round(_cell(row[C["eu_tot"]]) or 0.0, 3),
            "eupct": round(_cell(row[C["eu_pct"]]) or 0.0, 4),
        })
    if len(out) < 20:
        raise RuntimeError("feuille Kiel : %d pays seulement" % len(out))

    # Série mensuelle des engagements (feuille de la figure « Comm over Time »)
    monthly = []
    for sheet in ("Allocations by type and month", "Comm. by type and month"):
        if sheet not in wb.sheetnames:
            continue
        try:
            w2 = wb[sheet]
            rr = list(w2.iter_rows(values_only=True))
            hi = None
            for i, row in enumerate(rr[:30]):
                vals = [str(c or "").strip().lower() for c in (row or [])]
                if any(v in ("month", "date", "period") for v in vals):
                    hi = i
                    break
            if hi is None:
                continue
            hd = [str(c or "").strip().lower() for c in rr[hi]]
            ci = {}
            for i, h in enumerate(hd):
                for k, pat in (("mil", "military"), ("fin", "financial"), ("hum", "humanitarian")):
                    if h.startswith(pat):
                        ci[k] = i
                if h in ("month", "date", "period"):
                    ci["d"] = i
            if "d" not in ci or not ci.get("mil"):
                continue
            for row in rr[hi + 1:]:
                if not row or row[ci["d"]] is None:
                    continue
                d = row[ci["d"]]
                d = d.strftime("%Y-%m") if hasattr(d, "strftime") else str(d)[:7]
                # La feuille se termine par une ligne « Total » sans date : elle
                # entrerait dans la série comme un mois de 356 Md€.
                if not re.match(r"^(19|20)\d\d-(0[1-9]|1[0-2])$", d):
                    continue
                rec = {"d": d}
                for k in ("mil", "fin", "hum"):
                    if k in ci:
                        rec[k] = round(_cell(row[ci[k]]) or 0.0, 3)
                monthly.append(rec)
            if monthly:
                break
        except Exception as e:
            warn("aide · feuille « %s » : %s" % (sheet, e))

    wb.close()
    try:
        os.remove(tmp)
    except Exception:
        pass

    tot_all = round(sum(c["tot"] for c in out), 2)
    eu_side = round(sum(c["tot"] for c in out if c["geoeu"]), 2)
    us = next((c["tot"] for c in out if c["fr"] == "États-Unis"), None)
    log("aide · %d donateurs · %.1f Md€ alloués (Europe %.1f · US %s)"
        % (len(out), tot_all, eu_side, ("%.1f" % us) if us else "?"))
    return {
        "release": release, "url": url,
        "countries": sorted(out, key=lambda c: -c["tot"]),
        "monthly": monthly,
        "tot_alloc": tot_all, "eu_alloc": eu_side, "us_alloc": us,
    }


# ═══════════════════════════════════════════════════════════════════════════
# 3 · BANQUE NATIONALE D'UKRAINE
# ═══════════════════════════════════════════════════════════════════════════
def build_nbu(sess):
    out = {}
    # Réserves internationales, mensuel, en M USD.
    try:
        r = sess.get("https://bank.gov.ua/NBUStatService/v1/statdirectory/res?json")
        r.raise_for_status()
        rows = [x for x in r.json()
                if x.get("id_api") == "RES_OffReserveAssets" and x.get("value") is not None]
        ser = sorted({(str(x["dt"])[:4] + "-" + str(x["dt"])[4:6]): float(x["value"])
                      for x in rows}.items())
        out["reserves"] = [[d, round(v, 1)] for d, v in ser]
        log("nbu · réserves : %d mois, dernier %s = %.1f Md$"
            % (len(ser), ser[-1][0], ser[-1][1] / 1000.0))
    except Exception as e:
        warn("nbu · réserves : %s" % e)

    # Hryvnia / dollar, cours officiel, mensuel depuis 2021 (fin de mois).
    try:
        pts = []
        today = dt.date.today()
        y = 2021
        while y <= today.year:
            for m in range(1, 13):
                d = dt.date(y, m, 28)
                if d > today:
                    break
                pts.append(d)
            y += 1
        ser = []
        for d in pts:
            try:
                rr = sess.get("https://bank.gov.ua/NBUStatService/v1/statdirectory/"
                              "exchange?valcode=USD&date=%s&json" % d.strftime("%Y%m%d"))
                j = rr.json()
                if j and j[0].get("rate"):
                    ser.append([d.strftime("%Y-%m"), round(float(j[0]["rate"]), 4)])
            except Exception:
                pass
            time.sleep(0.05)
        if ser:
            out["uah"] = ser
            log("nbu · UAH/USD : %d mois, dernier %s = %.2f" % (len(ser), ser[-1][0], ser[-1][1]))
    except Exception as e:
        warn("nbu · hryvnia : %s" % e)
    return out or None


# ═══════════════════════════════════════════════════════════════════════════
# 4 · RÉFUGIÉS — HCR
# ═══════════════════════════════════════════════════════════════════════════
def build_unhcr(sess):
    base = ("https://api.unhcr.org/population/v1/population/"
            "?yearFrom=2014&yearTo=%d&coo=UKR&coa_all=false&limit=1000&page=%d")
    yr = dt.date.today().year
    tot = {}
    page = 1
    while page <= 25:
        r = sess.get(base % (yr, page))
        r.raise_for_status()
        j = r.json()
        items = j.get("items") or []
        if not items:
            break
        for it in items:
            y = int(it.get("year") or 0)
            if not y:
                continue
            a = tot.setdefault(y, {"ref": 0, "asy": 0, "idp": 0})
            for k, f in (("ref", "refugees"), ("asy", "asylum_seekers"), ("idp", "idps")):
                try:
                    a[k] += int(float(it.get(f) or 0))
                except Exception:
                    pass
        if page >= int(j.get("maxPages") or 1):
            break
        page += 1
    if not tot:
        raise RuntimeError("HCR : aucune ligne")
    ser = [{"y": y, "ref": v["ref"], "asy": v["asy"], "idp": v["idp"]}
           for y, v in sorted(tot.items())]
    log("unhcr · %d années · %d réfugiés ukrainiens en %d"
        % (len(ser), ser[-1]["ref"], ser[-1]["y"]))
    return {"series": ser}


# ═══════════════════════════════════════════════════════════════════════════
# 5 · DÉCOUPLAGE ÉNERGÉTIQUE — Eurostat
# ═══════════════════════════════════════════════════════════════════════════
def estat(sess, ds, params):
    q = "&".join("%s=%s" % (k, v) for k, v in params)
    r = sess.get("%s%s?format=JSON&lang=EN&%s" % (EUROSTAT, ds, q))
    r.raise_for_status()
    j = r.json()
    dim = j.get("dimension") or {}
    tix = ((dim.get("time") or {}).get("category") or {}).get("index") or {}
    inv = {int(v): k for k, v in tix.items()}
    n_time = len(inv)
    vals = j.get("value") or {}
    out = {}
    for k, v in vals.items():
        # l'index est plat : la dimension temps est la dernière → modulo
        t = inv.get(int(k) % n_time)
        if t and v is not None:
            out[t] = float(v)
    return sorted(out.items())


def build_eurostat(sess):
    out = {}
    try:
        # Importations UE de gaz naturel en provenance de Russie (TJ PCS)
        g = estat(sess, "nrg_ti_gasm",
                  [("partner", "RU"), ("geo", "EU27_2020"), ("siec", "G3000"),
                   ("unit", "MIO_M3"), ("sinceTimePeriod", "2019-01")])
        if g:
            out["gas_ru"] = [[d, round(v, 1)] for d, v in g]
            log("eurostat · gaz RU→UE : %d mois, dernier %s = %.0f Mm³" % (len(g), g[-1][0], g[-1][1]))
    except Exception as e:
        warn("eurostat · gaz : %s" % e)
    try:
        o = estat(sess, "nrg_ti_oilm",
                  [("partner", "RU"), ("geo", "EU27_2020"), ("siec", "O4100_TOT"),
                   ("unit", "THS_T"), ("sinceTimePeriod", "2019-01")])
        if o:
            out["oil_ru"] = [[d, round(v, 1)] for d, v in o]
            log("eurostat · pétrole RU→UE : %d mois, dernier %s = %.0f kt" % (len(o), o[-1][0], o[-1][1]))
    except Exception as e:
        warn("eurostat · pétrole : %s" % e)
    try:
        gt = estat(sess, "nrg_ti_gasm",
                   [("partner", "TOTAL"), ("geo", "EU27_2020"), ("siec", "G3000"),
                    ("unit", "MIO_M3"), ("sinceTimePeriod", "2019-01")])
        if gt:
            out["gas_tot"] = [[d, round(v, 1)] for d, v in gt]
    except Exception as e:
        warn("eurostat · gaz total : %s" % e)
    return out or None


# ═══════════════════════════════════════════════════════════════════════════
# 6 · MER NOIRE — IMF PortWatch (escales quotidiennes, AIS)
# ═══════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════
# 6 · COMMERCE EXTÉRIEUR — Eurostat, vu depuis l'Union européenne
#
# ⚠ SOURCE ESSAYÉE PUIS ÉCARTÉE (2026-08-28). Ce bloc devait porter les escales
# des ports ukrainiens relevées par AIS satellitaire (IMF PortWatch) : le
# corridor céréalier de la mer Noire s'y lit à l'œil nu. Vérification faite
# port par port AVANT publication, la série ukrainienne de PortWatch est
# inexploitable à partir de 2025 — le trafic y est attribué aux mauvais ports.
# Relevé, en escales annuelles :
#
#     Odessa       2024 :   743   ->  2025 :     4
#     Chornomorsk  2024 :   893   ->  2025 :     2
#     Pivdennyi    2024 :   507   ->  2025 :     3
#     Kertch       2024 :   382   ->  2025 : 3 286
#
# Odessa n'est pas à l'arrêt : c'est par elle que passe l'essentiel des
# exportations ukrainiennes depuis l'ouverture du corridor. Un petit port de
# Crimée qui capterait huit fois le trafic du premier port du pays est une
# erreur d'affectation en amont, pas un fait de guerre. Publier ce graphique
# aurait affirmé que les ports ukrainiens sont morts.
#
# On prend donc la même question — « le commerce ukrainien tient-il ? » — par
# une source qui, elle, tient : les DOUANES des Vingt-Sept. Elle ne voit qu'un
# partenaire, mais l'UE pèse aujourd'hui la majeure partie du commerce extérieur
# ukrainien, et personne ne conteste un dédouanement européen.
# ═══════════════════════════════════════════════════════════════════════════
TRADE_DS = "ext_st_eu27_2020sitc"


def build_trade(sess):
    out = {}
    for key, flow, lab in (("imp", "IMP", "importations de l'UE en provenance d'Ukraine"),
                           ("exp", "EXP", "exportations de l'UE vers l'Ukraine")):
        try:
            ser = estat(sess, TRADE_DS,
                        [("partner", "UA"), ("geo", "EU27_2020"), ("sitc06", "TOTAL"),
                         ("indic_et", "TRD_VAL"), ("stk_flow", flow),
                         ("sinceTimePeriod", "2019-01")])
            if ser:
                out[key] = [[d, round(v, 1)] for d, v in ser]
                log("commerce · %s : %d mois, dernier %s = %.0f M€"
                    % (lab, len(ser), ser[-1][0], ser[-1][1]))
        except Exception as e:
            warn("commerce · %s : %s" % (lab, e))
    if not out:
        raise RuntimeError("Eurostat : aucune série de commerce UE↔Ukraine")
    out["refused"] = {
        "source": "IMF PortWatch",
        "why": ("Les escales des ports ukrainiens ont été écartées : à partir de 2025, "
                "la série attribue le trafic aux mauvais ports — Odessa passe de 743 "
                "escales en 2024 à 4 en 2025, tandis que Kertch, petit port de Crimée, "
                "monte de 382 à 3 286. Odessa n'est pas à l'arrêt : c'est par elle que "
                "passe l'essentiel des exportations depuis l'ouverture du corridor. Le "
                "corridor de la mer Noire n'est donc pas représenté ici tant que la "
                "source ne redevient pas cohérente."),
    }
    return out


# ═══════════════════════════════════════════════════════════════════════════
# Repères documentés (pas de collecte : des rapports, cités et datés)
# ═══════════════════════════════════════════════════════════════════════════
REFERENCES = [
    {
        "k": "rdna",
        "label": "Besoins de reconstruction et de relèvement",
        "value": 524.0, "unit": "Md$", "asof": "2025-02",
        "src": "Banque mondiale · Gouvernement d'Ukraine · Commission européenne · ONU — RDNA4",
        "url": "https://www.worldbank.org/en/news/press-release/2025/02/25/updated-ukraine-recovery-needs-assessment",
        "note": "Quatrième évaluation rapide des dommages et des besoins (RDNA4), arrêtée "
                "au 31 décembre 2024. Dommages physiques directs recensés : 176 Md$.",
    },
    {
        "k": "damage",
        "label": "Dommages physiques directs recensés",
        "value": 176.0, "unit": "Md$", "asof": "2025-02",
        "src": "Banque mondiale — RDNA4",
        "url": "https://www.worldbank.org/en/news/press-release/2025/02/25/updated-ukraine-recovery-needs-assessment",
        "note": "Logement, transports, énergie et commerce concentrent l'essentiel.",
    },
    {
        "k": "frozen",
        "label": "Avoirs de la banque centrale russe immobilisés",
        "value": 210.0, "unit": "Md€", "asof": "2024-06",
        "src": "Conseil de l'Union européenne / G7",
        "url": "https://www.consilium.europa.eu/en/policies/sanctions-against-russia/",
        "note": "Dont environ 185 Md€ chez Euroclear en Belgique. Les avoirs sont "
                "immobilisés, pas confisqués : seuls leurs revenus sont mobilisés.",
    },
]


# ═══════════════════════════════════════════════════════════════════════════
def load_prev():
    for base in CACHES:
        p = base + ".json"
        if os.path.exists(p):
            try:
                with open(p, encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
    return None


def write_out(out):
    blob = json.dumps(out, ensure_ascii=False, separators=(",", ":"))
    hdr = "/* ukraine_cache.js — Guerre en Ukraine (DeepStateMap · IfW Kiel · NBU · HCR · Eurostat) */"
    wrote = 0
    for base in CACHES:
        try:
            os.makedirs(os.path.dirname(base), exist_ok=True)
            for ext, txt in ((".json", blob),
                             (".js", hdr + "\nwindow.__ATLAS_UKRAINE__ = " + blob + ";\n")):
                tmp = base + ext + ".tmp"
                with open(tmp, "w", encoding="utf-8") as f:
                    f.write(txt)
                os.replace(tmp, base + ext)
            log("écrit %s.json (%d Ko)" % (base, len(blob) // 1024))
            wrote += 1
        except Exception as e:
            warn("écriture %s : %s" % (base, e))
    if not wrote:
        raise RuntimeError("aucune copie écrite")


def main():
    mode = "live"
    for a in sys.argv[1:]:
        if a in ("--full", "--live", "--front"):
            mode = a[2:]
    prev = load_prev() or {}
    sess = _sess()
    out = {"meta": {}, "front": None, "aid": None, "nbu": None,
           "unhcr": None, "eurostat": None, "trade": None, "refs": REFERENCES}
    status = {}

    def run(key, fn, keep_prev=True):
        t0 = time.time()
        try:
            out[key] = fn()
            status[key] = {"ok": 1, "s": round(time.time() - t0, 1)}
        except Exception as e:
            warn("%s : %s" % (key, e))
            status[key] = {"ok": 0, "err": str(e)[:200]}
            if keep_prev and prev.get(key):
                out[key] = prev[key]
                status[key]["reused"] = 1

    run("front", lambda: build_front(sess, prev, full=(mode == "full")))
    if mode != "front":
        # Blocs lents : en --live on ne les refait que si le cache a vieilli.
        pm = (prev.get("meta") or {})
        age_h = (time.time() - float(pm.get("slow_ts") or 0)) / 3600.0
        redo_slow = (mode == "full") or age_h > MAXAGE_H or not prev.get("aid")
        if redo_slow:
            run("aid", lambda: build_aid(sess))
            run("nbu", lambda: build_nbu(sess))
            run("unhcr", lambda: build_unhcr(sess))
            run("eurostat", lambda: build_eurostat(sess))
            run("trade", lambda: build_trade(sess))
            slow_ts = int(time.time())
        else:
            log("blocs lents repris du cache (%.1f h)" % age_h)
            for k in ("aid", "nbu", "unhcr", "eurostat", "trade"):
                out[k] = prev.get(k)
                status[k] = {"ok": 1, "cached": 1}
            slow_ts = int(float(pm.get("slow_ts") or 0))
    else:
        for k in ("aid", "nbu", "unhcr", "eurostat", "trade"):
            out[k] = prev.get(k)
        slow_ts = int(float((prev.get("meta") or {}).get("slow_ts") or 0))

    if not out.get("front"):
        sys.stderr.write("ÉCHEC : bloc front indisponible et aucun cache antérieur.\n")
        sys.exit(1)

    now = time.time()
    out["meta"] = {
        "built_at": dt.datetime.fromtimestamp(now, dt.timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "built_ts": int(now),
        "slow_ts": slow_ts,
        "mode": mode,
        "version": 1,
        "status": status,
        "warnings": WARN[:20],
        "n_front": len((out["front"] or {}).get("series") or []),
    }
    out["sources"] = [
        {"id": "dsm", "label": "DeepStateMap", "url": "https://deepstatemap.live/",
         "note": "Tracé du front publié en GeoJSON. Les surfaces affichées ici ne sont "
                 "pas reprises d'un communiqué : elles sont calculées à partir de ce "
                 "tracé, par excès sphérique, et contrôlées sur la Crimée."},
        {"id": "kiel", "label": "IfW Kiel — Ukraine Support Tracker",
         "url": "https://www.ifw-kiel.de/topics/war-against-ukraine/ukraine-support-tracker/",
         "note": "Aide bilatérale engagée et allouée à l'Ukraine, par donateur."},
        {"id": "nbu", "label": "Banque nationale d'Ukraine", "url": "https://bank.gov.ua/",
         "note": "Cours officiel de la hryvnia et réserves internationales."},
        {"id": "unhcr", "label": "HCR — Nations unies", "url": "https://www.unhcr.org/refugee-statistics/",
         "note": "Réfugiés d'origine ukrainienne, par pays d'accueil."},
        {"id": "estat", "label": "Eurostat", "url": "https://ec.europa.eu/eurostat/",
         "note": "Importations de gaz et de pétrole de l'UE en provenance de Russie."},
        {"id": "estat_trade", "label": "Eurostat — commerce extérieur",
         "url": "https://ec.europa.eu/eurostat/databrowser/view/ext_st_eu27_2020sitc/",
         "note": "Échanges de marchandises entre l'UE et l'Ukraine, mensuel, "
                 "déclarés par les douanes des Vingt-Sept."},
        {"id": "wb", "label": "Banque mondiale — RDNA", "url": "https://www.worldbank.org/",
         "note": "Évaluation des dommages et des besoins de reconstruction."},
    ]
    write_out(out)
    f = out["front"]["last"]
    log("OK — front %s : %s km² sous contrôle russe (%.2f %% du territoire) · "
        "%d points d'historique · %d avertissement(s)"
        % (f["d"], format(int(f["tot"]), ",d").replace(",", " "), f["pct"],
           out["meta"]["n_front"], len(WARN)))


if __name__ == "__main__":
    main()
