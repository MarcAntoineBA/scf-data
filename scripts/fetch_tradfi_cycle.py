#!/usr/bin/env python3
"""fetch_tradfi_cycle.py — Données de l'onglet "Cycle Tracker · TradFi".

Cyclicité séculaire ACTIFS FINANCIERS ↔ ACTIFS RÉELS (or & matières premières) :
le grand pendule intermarché (thèses Dow/Gold, supercycle des matières premières,
real assets vs financial assets). Pendant TradFi du Cycle Tracker BTC.

Indicateurs (historique maximal, tout gratuit) :
  • S&P 500 / Or .......... datahub Shiller (S&P 1871+) ÷ datahub gold (1833+)  [maître]
  • Matières prem. / Actions  FRED PPIACO (1913+) ÷ S&P 500                      [supercycle]
  • Or réel ............... gold ÷ CPI (Shiller, 1833+)
  • Indice MP réel ........ FRED PPIACO ÷ CPI (1913+)
  • Or / Argent ........... gold ÷ silver (Yahoo, 2000+)
  • Cuivre / Or ........... copper ÷ gold (Yahoo, 2000+)

Composite "Pendule financier↔réel" 0-100 : 100 = actions au sommet / réel délaissé
(1966, 1999, 2021) ; 0 = or & matières premières au sommet (1980, 2011). Construit
par PERCENTILE historique (auditable, pas de "juste valeur" arbitraire).

Sources : datahub.io (CSV GitHub stables), FRED API (clé _fred_helpers), Yahoo v8
(curl_cffi). Splice : history mensuelle longue + dernier mois Yahoo pour le live.

Usage : python3 fetch_tradfi_cycle.py [--force]
"""

import bisect
import csv
import io
import json
import math
import os
import sys
import time
import urllib.request
from datetime import datetime, timezone, date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path.home() / "Desktop" / "Site_Crypto_Finance"))
try:
    from _fred_helpers import fetch_fred
except Exception:
    fetch_fred = None

# ─────────────────────────── Config ────────────────────────────
CACHE_DIR = Path.home() / "Library" / "Caches" / "site_crypto_finance"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
CACHE_JSON = CACHE_DIR / "tradfi_cycle_cache.json"
CACHE_JS = CACHE_DIR / "tradfi_cycle_cache.js"
SITE_DIR = Path.home() / "Desktop" / "Site_Crypto_Finance"
CACHE_MAX_HOURS = 12

GOLD_URL = "https://raw.githubusercontent.com/datasets/gold-prices/master/data/monthly.csv"
SHILLER_URL = "https://raw.githubusercontent.com/datasets/s-and-p-500/master/data/data.csv"
# World Bank « Pink Sheet » : prix matières premières mensuels depuis 1960 (or/argent/cuivre).
# Seule source longue & gratuite pour argent & cuivre — les futures Yahoo ne remontent qu'à ~2000.
WB_PINKSHEET_URLS = [
    "https://thedocs.worldbank.org/en/doc/5d903e848db1d1b83e0ec8f744e55570-0350012021/related/CMO-Historical-Data-Monthly.xlsx",
    "https://www.worldbank.org/content/dam/sites/commodity-markets/data/CMO-Historical-Data-Monthly.xlsx",
]
UA = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) CapitalAntifragile/1.0"}

COL = {"red": "#ff5570", "orange": "#f5a623", "green": "#28e07e", "bottom": "#1ec5ff"}


def log(m):
    sys.stderr.write(f"[tradfi-cycle] {m}\n")


# ─────────────────────── Fetch helpers ─────────────────────────
def http_get(url, timeout=40):
    req = urllib.request.Request(url, headers=UA)
    return urllib.request.urlopen(req, timeout=timeout).read().decode("utf-8", "replace")


def gold_monthly():
    """Or mensuel USD/oz, 1833+ (datahub / LBMA). {YYYY-MM: prix}."""
    out = {}
    for r in csv.DictReader(io.StringIO(http_get(GOLD_URL))):
        d, p = r.get("Date", ""), r.get("Price", "")
        if d and p:
            try:
                out[d[:7]] = float(p)
            except ValueError:
                pass
    if len(out) < 500:
        raise RuntimeError(f"gold datahub: {len(out)} pts")
    return out


def worldbank_monthly():
    """World Bank « Pink Sheet » : prix mensuels depuis 1960 pour or/argent/cuivre.
    Retourne {'gold':{YYYY-MM:px}, 'silver':{...}, 'copper':{...}} (vide si indisponible).
    Donne à l'argent & au cuivre l'historique 1960+ que les futures Yahoo (~2000) n'ont pas."""
    try:
        import openpyxl
    except Exception as e:
        log(f"WARN WorldBank: openpyxl absent ({e})")
        return {}
    raw = None
    for u in WB_PINKSHEET_URLS:
        try:
            req = urllib.request.Request(u, headers=UA)
            raw = urllib.request.urlopen(req, timeout=60).read()
            break
        except Exception as e:
            log(f"WARN WorldBank {u[:48]}… : {e}")
    if not raw:
        return {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb["Monthly Prices"] if "Monthly Prices" in wb.sheetnames else wb[wb.sheetnames[0]]
        want = {"Gold": "gold", "Silver": "silver", "Copper": "copper"}
        colmap, out = {}, {"gold": {}, "silver": {}, "copper": {}}
        names = None
        for r in ws.iter_rows(values_only=True):
            if names is None:
                if r and any(c == "Gold" for c in r if isinstance(c, str)):
                    names = r
                    for i, c in enumerate(r):
                        if c in want:
                            colmap[want[c]] = i
                continue
            d = r[0]
            if not isinstance(d, str) or "M" not in d:
                continue
            ym = d[:4] + "-" + d[5:7]        # '1960M01' → '1960-01'
            for key, ci in colmap.items():
                try:
                    fv = float(r[ci])
                    if fv > 0:
                        out[key][ym] = fv
                except (TypeError, ValueError):
                    pass
        n = {k: len(v) for k, v in out.items()}
        log(f"WorldBank Pink Sheet : or {n['gold']} · argent {n['silver']} · cuivre {n['copper']}")
        if n["silver"] < 300 or n["copper"] < 300:
            return {}
        return out
    except Exception as e:
        log(f"WARN WorldBank parse : {e}")
        return {}


def shiller_monthly():
    """S&P 500 + CAPE(PE10) + CPI mensuels, 1871+ (datahub / Shiller)."""
    sp, cape, cpi = {}, {}, {}
    for r in csv.DictReader(io.StringIO(http_get(SHILLER_URL))):
        d = (r.get("Date") or "")[:7]
        if not d:
            continue

        def f(key):
            v = r.get(key, "")
            try:
                x = float(v)
                return x if x > 0 else None
            except (ValueError, TypeError):
                return None
        v = f("SP500")
        if v:
            sp[d] = v
        c = f("PE10")
        if c:
            cape[d] = c
        p = f("Consumer Price Index")
        if p:
            cpi[d] = p
    if len(sp) < 500:
        raise RuntimeError(f"shiller S&P: {len(sp)} pts")
    return sp, cape, cpi


def fred_monthly(series_id):
    """Série FRED → {YYYY-MM: valeur} (dernière valeur du mois)."""
    if fetch_fred is None:
        return {}
    try:
        r = fetch_fred(series_id, start="1900-01-01")
    except Exception as e:
        log(f"WARN FRED {series_id}: {e}")
        return {}
    out = {}
    if r and r.get("dates"):
        for d, v in zip(r["dates"], r["values"]):
            if v is not None:
                out[d[:7]] = float(v)
    return out


def _yahoo(symbol, interval, fmt, daily_since=None):
    try:
        from curl_cffi import requests as cr
        if daily_since is not None:  # vrai daily via period1/period2 (range=max sous-échantillonne)
            p1 = int(datetime(daily_since, 1, 1, tzinfo=timezone.utc).timestamp())
            p2 = int(time.time())
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?period1={p1}&period2={p2}&interval={interval}"
        else:
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?range=max&interval={interval}"
        j = cr.get(url, impersonate="chrome120", timeout=30).json()
        res = j["chart"]["result"][0]
        ts, cl = res["timestamp"], res["indicators"]["quote"][0]["close"]
        out = {}
        for t, c in zip(ts, cl):
            if c is None:
                continue
            out[datetime.fromtimestamp(t, timezone.utc).strftime(fmt)] = float(c)
        return out
    except Exception as e:
        log(f"WARN Yahoo {symbol} ({interval}): {e}")
        return {}


def yahoo_monthly(symbol):
    """Clôtures mensuelles Yahoo v8 (curl_cffi). {YYYY-MM: close}."""
    return _yahoo(symbol, "1mo", "%Y-%m")


def yahoo_daily(symbol, since=2000):
    """Clôtures quotidiennes Yahoo v8 (curl_cffi, vraie résolution daily). {YYYY-MM-DD: close}."""
    return _yahoo(symbol, "1d", "%Y-%m-%d", daily_since=since)


def splice(base, extra):
    """Complète `base` (history longue) avec les mois de `extra` (Yahoo récent)
    absents de base — pour amener le 'now' à aujourd'hui sans réécrire l'histoire."""
    out = dict(base)
    if base:
        last = max(base)
        for k, v in extra.items():
            if k > last:
                out[k] = v
    return out


# ─────────────────────── Math helpers ──────────────────────────
def ym_to_unix(ym):
    return int(datetime(int(ym[:4]), int(ym[5:7]), 1, tzinfo=timezone.utc).timestamp())


def ratio(num, den):
    """Ratio aligné sur les mois communs. {YYYY-MM: num/den}."""
    out = {}
    for k in num.keys() & den.keys():
        if den[k]:
            out[k] = num[k] / den[k]
    return out


def to_series(d):
    """{YYYY-MM: val} → [[unixTs, round(val)]] trié, pour Chart.js."""
    return [[ym_to_unix(k), round(d[k], 6)] for k in sorted(d)]


def ymd_to_unix(d):
    return int(datetime(int(d[:4]), int(d[5:7]), int(d[8:10]), tzinfo=timezone.utc).timestamp())


def to_series_d(d):
    """{YYYY-MM-DD: val} → [[unixTs, round(val)]] trié (séries quotidiennes)."""
    return [[ymd_to_unix(k), round(d[k], 6)] for k in sorted(d)]


def ratio_keys(num, den):
    """Ratio sur clés communes (fonctionne pour YYYY-MM ou YYYY-MM-DD)."""
    out = {}
    for k in num.keys() & den.keys():
        if den[k]:
            out[k] = num[k] / den[k]
    return out


def pct_rank(values, v):
    """Percentile (0-100) de v dans values."""
    if not values:
        return None
    n = sum(1 for x in values if x <= v)
    return 100.0 * n / len(values)


def last_val(d):
    return d[max(d)] if d else None


def monthly_returns(price):
    """{YYYY-MM: prix} → matrice rendements % {année: {mois: pct}} (mensuel close-to-close)."""
    keys = sorted(price)
    out = {}
    for i in range(1, len(keys)):
        prev, cur = price[keys[i - 1]], price[keys[i]]
        if prev and prev > 0:
            y, m = keys[i][:4], int(keys[i][5:7])
            out.setdefault(y, {})[str(m)] = round((cur / prev - 1) * 100, 2)
    return out


# ─────────────────────── Zones / scoring ───────────────────────
# Toutes les jauges sont orientées "PENDULE" : score haut = actifs FINANCIERS
# dominants / actifs réels bon marché (sommet actions, ex. 1999) ; score bas =
# actifs RÉELS dominants (or & MP au sommet, ex. 1980).
def zone_from_score(s):
    if s is None:
        return None
    if s >= 75:
        return "red"      # actions euphoriques / réel délaissé (sommet du cycle financier)
    if s >= 50:
        return "orange"
    if s >= 25:
        return "green"
    return "bottom"        # or & matières premières au sommet


ZONE_LABEL = {
    "red": "Actions au sommet · réel délaissé",
    "orange": "Côté financier",
    "green": "Bascule vers le réel",
    "bottom": "Or & MP au sommet",
}


# ─────────────────────────── Build ─────────────────────────────
def build():
    # 1) Séries brutes
    gold = gold_monthly()
    # Correction étalon-or : la source (datahub) est inexacte avant 1934 (~$18.9 au lieu
    # du prix OFFICIEL US). On recale sur les prix légaux documentés du Trésor américain :
    #   $19.39/oz (Coinage Act 1792-1834) · $20.67/oz (1834-1933) · puis $35 dès 1934 (Gold
    #   Reserve Act, conservé via datahub qui est exact à partir de là). Le marché LBMA prend
    #   le relais après 1968 (datahub) puis Yahoo daily 2000+.
    n_fix = 0
    for k in list(gold):
        if k < "1834-01":
            gold[k] = 19.39; n_fix += 1
        elif k < "1934-01":
            gold[k] = 20.67; n_fix += 1
    log(f"or : recalage étalon-or sur {n_fix} mois (prix officiels US pré-1934)")
    sp, cape, cpi = shiller_monthly()
    ppi = fred_monthly("PPIACO")          # commodities PPI, 1913+
    cpi_fred = fred_monthly("CPIAUCNS")   # déflateur 1913+ (complément Shiller)
    log(f"gold {len(gold)} ({min(gold)}→{max(gold)}) · S&P {len(sp)} · CAPE {len(cape)} · PPI {len(ppi)}")

    # Live mensuel : amener or & S&P au mois courant via Yahoo
    gold_m = splice(gold, yahoo_monthly("GC=F"))
    sp = splice(sp, yahoo_monthly("%5EGSPC"))
    # Argent & cuivre : World Bank Pink Sheet (1960+) comme base longue, Yahoo futures pour le live.
    # Repli sur Yahoo seul (~2000+) si la Banque Mondiale est indisponible.
    wb_px = worldbank_monthly()
    yh_silver, yh_copper = yahoo_monthly("SI=F"), yahoo_monthly("HG=F")
    silver_m = splice(wb_px["silver"], yh_silver) if wb_px.get("silver") else yh_silver
    copper_m = splice(wb_px["copper"], yh_copper) if wb_px.get("copper") else yh_copper
    log(f"argent {len(silver_m)} ({min(silver_m)}→{max(silver_m)}) · "
        f"cuivre {len(copper_m)} ({min(copper_m)}→{max(copper_m)})")
    # CPI : Shiller + FRED CPIAUCNS pour le live
    cpi = splice(cpi, cpi_fred)

    # DAILY (précis & live) pour les prix/ratios métaux — futures Yahoo (~2000+)
    gold_d = yahoo_daily("GC=F")
    silver_d = yahoo_daily("SI=F")
    copper_d = yahoo_daily("HG=F")
    log(f"daily : or {len(gold_d)} · argent {len(silver_d)} · cuivre {len(copper_d)}")

    # 2) Ratios séculaires (mensuels, longue histoire) — S&P 500 comme référence (pas le Dow)
    sp_gold = ratio(sp, gold_m)             # S&P 500 / Or  (maître)
    comm_eq = ratio(ppi, sp)                # Matières premières / Actions
    comm_real = ratio(ppi, cpi)             # Indice MP réel (PPI ÷ CPI)
    # Rebase 1982=100 (MÊME base que le PPIACO nominal FRED, qui est 1982=100) → comparabilité
    # directe : en 1982 nominal & réel valent 100 ; aujourd'hui PPIACO≈284 mais réel≈82, l'écart EST
    # l'inflation cumulée. Scaling linéaire positif → percentile/zone du pendule inchangés.
    def rebase_1982(d, name):
        b = [v for k, v in d.items() if k.startswith("1982-")]
        if b:
            m = sum(b) / len(b)
            if m:
                return {k: v / m * 100.0 for k, v in d.items()}
        log(f"⚠ rebase {name}: pas de point 1982, serie brute conservee")
        return d
    comm_real = rebase_1982(comm_real, "comm_real")
    # Ratios métaux : DAILY (précis/live)
    gold_silver = ratio_keys(gold_d, silver_d)   # Or / Argent
    copper_gold = ratio_keys(copper_d, gold_d)   # Cuivre / Or

    # 3) Composite PENDULE (percentiles historiques, séries longues & robustes)
    def pctl(d):
        v = last_val(d)
        return (v, pct_rank(list(d.values()), v)) if d else (None, None)
    sg_now, sg_pct = pctl(sp_gold)
    ce_now, ce_pct = pctl(comm_eq)
    cr_now, cr_pct = pctl(comm_real)
    # 2 DIMENSIONS DÉCORRÉLÉES (évite le double comptage du PPI) :
    #   • Actions vs Or  : S&P/Gold haut = actions chères vs or = financier dominant
    #   • MP vs Actions  : commodities/actions BAS = MP bon marché = financier dominant (invert)
    parts = []
    if sg_pct is not None:
        parts.append(sg_pct)
    if ce_pct is not None:
        parts.append(100 - ce_pct)
    composite_val = round(sum(parts) / len(parts)) if parts else None
    cstatus = zone_from_score(composite_val)

    # 4) Table auditable (chaque jauge orientée pendule via son percentile)
    def gauge_score(d, invert=False):
        v = last_val(d)
        if v is None:
            return None, None, None
        p = pct_rank(list(d.values()), v)
        score = round(100 - p if invert else p)
        return v, round(p, 1), score

    def row(key, label, d, invert, fmt, formula, source):
        v, p, score = gauge_score(d, invert)
        st = zone_from_score(score)
        return {
            "key": key, "label": label,
            "value": round(v, 4) if v is not None else None,
            "display": (fmt(v) if v is not None else "—"),
            "score": score, "percentile": p, "status": st,
            "zone": ZONE_LABEL.get(st),
            "thresholds": "percentile historique → pendule : ≥75 actions/sommet · ≥50 financier · ≥25 bascule · <25 réel/sommet",
            "formula": formula, "source_url": source,
            "as_of": max(d) if d else None,
        }

    table = [
        row("sp_gold", "S&P 500 / Or", sp_gold, False,
            lambda v: f"{v:.2f}",
            "S&P 500 ÷ prix de l'or (oz) — le pendule maître actions vs or (équivalent long du Dow/Gold)",
            "https://www.longtermtrends.com/dow-gold-ratio/"),
        row("comm_eq", "Matières premières / Actions", comm_eq, True,
            lambda v: f"{v*1000:.1f}‰",
            "Indice prix matières premières (PPI) ÷ S&P 500 — bas = MP bon marché (début de supercycle)",
            "https://www.longtermtrends.com/stocks-commodities-ratio/"),
        row("comm_real", "Indice matières premières réel", comm_real, True,
            lambda v: f"{v:.0f}",
            "Indice prix matières premières (PPIACO) ÷ CPI, rebasé 1982=100 (même base que le PPIACO nominal). "
            "Mesure le pouvoir d'achat réel des MP : nominal 1982=100→284 aujourd'hui, réel 1982=100→~82 (l'écart = inflation). Bas = MP sous-évaluées en réel",
            "https://fred.stlouisfed.org/series/PPIACO"),
    ]

    # 5) Composite breakdown (dimensions : actions↔or, matières prem.↔actions)
    composite = {
        "value": composite_val, "status": cstatus,
        "label": {"red": "Actions au sommet — réel historiquement délaissé",
                  "orange": "Côté actifs financiers, mais en bascule",
                  "green": "Rotation enclenchée vers les actifs réels",
                  "bottom": "Or & matières premières au sommet du cycle"}.get(cstatus, "—"),
        "dimensions": [
            {"key": "sp_gold", "label": "Actions vs Or (S&P 500 / Or)",
             "value": round(sg_pct) if sg_pct is not None else None,
             "status": zone_from_score(sg_pct), "display": f"{sg_now:.2f}" if sg_now else "—"},
            {"key": "comm_eq", "label": "Matières premières vs Actions",
             "value": round(100 - ce_pct) if ce_pct is not None else None,
             "status": zone_from_score(100 - ce_pct) if ce_pct is not None else None,
             "display": f"{ce_now*1000:.1f}‰" if ce_now else "—"},
        ],
        "scores": [round(p) for p in parts],
        "formula": "Pendule 0-100 par PERCENTILE historique, 2 dimensions DÉCORRÉLÉES : "
                   "Actions vs Or (S&P 500 / Or — haut = actions chères vs or) + Matières premières vs "
                   "Actions (PPI / S&P 500, inversé — bas = MP bon marché). Moyenne des deux. "
                   "100 = euphorie actions / actifs réels au plancher (1966, 1999, 2021) · "
                   "0 = or & matières premières au sommet du cycle (1980, 2011).",
    }

    # 5b) Historique du PENDULE (composite reconstruit, même formule que la jauge live :
    #     percentile statique de chaque dimension → moyenne des dispos). Dernier point = valeur live.
    sg_sorted = sorted(sp_gold.values())
    ce_sorted = sorted(comm_eq.values())
    def srank(sv, v):
        return 100.0 * bisect.bisect_right(sv, v) / len(sv)
    ce_max_k = max(comm_eq) if comm_eq else None
    ce_last_inv = (100 - srank(ce_sorted, comm_eq[ce_max_k])) if ce_max_k else None
    pend = {}
    for k in sorted(set(sp_gold) | set(comm_eq)):
        ps = []
        if k in sp_gold:
            ps.append(srank(sg_sorted, sp_gold[k]))
        if k in comm_eq:
            ps.append(100 - srank(ce_sorted, comm_eq[k]))
        elif ce_max_k and k > ce_max_k and ce_last_inv is not None:
            ps.append(ce_last_inv)   # forward-fill comm_eq (PPI retarde) → dernier point = valeur live
        if ps:
            pend[k] = sum(ps) / len(ps)
    pendule_hist = to_series(pend)

    # 6) Saisonnalité (rendements mensuels) — Or, Argent, Cuivre, S&P 500
    seasonality = {
        "gold": monthly_returns(gold_m),
        "silver": monthly_returns(silver_m),
        "copper": monthly_returns(copper_m),
        "sp500": monthly_returns(sp),
    }

    # 7) Comparaison des cycles séculaires : S&P/Or normalisé depuis chaque sommet
    #    (les grandes rotations financier→réel). x = mois depuis le sommet.
    peaks = [("1966-01", "Cycle 1966 → 1980"), ("1999-08", "Cycle 1999 → 2011"),
             ("2021-12", "Cycle 2021 → … (en cours)")]
    sg_keys = sorted(sp_gold)
    secular = []
    for pk, lbl in peaks:
        if pk not in sp_gold:
            # plus proche mois disponible
            near = [k for k in sg_keys if k >= pk]
            if not near:
                continue
            pk = near[0]
        base = sp_gold[pk]
        pts = []
        for k in sg_keys:
            if k < pk:
                continue
            months = (int(k[:4]) - int(pk[:4])) * 12 + (int(k[5:7]) - int(pk[5:7]))
            if months > 240:
                break
            pts.append([months, round(sp_gold[k] / base, 4)])
        secular.append({"label": lbl, "peak": pk, "data": pts})

    # Or nominal : affiché depuis 1871, même base de départ que le S&P 500 et le Pendule.
    # Avant août 1971 le prix est ADMINISTRÉ par la loi (étalon-or : $19.39 → $20.67 → $35),
    # il n'est pas coté : la courbe est un escalier plat. Le front-end trace cette portion en
    # pointillés atténués pour ne pas la faire passer pour une cotation de marché.
    # Mensuel 1871→2000 + DAILY Yahoo 2000+ (précis & live).
    GOLD_DISPLAY_START = "1871-01"
    daily_cut = (min(gold_d)[:7] if gold_d else "2000-09")   # mensuel jusqu'au 1er mois daily (pas de trou)
    gold_nom_series = [[ym_to_unix(k), round(gold_m[k], 4)] for k in sorted(gold_m) if GOLD_DISPLAY_START <= k < daily_cut]
    gold_nom_series += to_series_d(gold_d)
    gold_nom_series.sort(key=lambda r: r[0])

    payload = {
        "meta": {
            "updated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
            "updated_unix": int(time.time()),
            "history_start": min(min(sp_gold), min(comm_real)) if sp_gold else None,
            "sources": {
                "gold": "Prix officiels US (étalon-or $19.39/$20.67 pré-1934) + datahub/LBMA (1934+) + Yahoo GC=F daily (2000+)",
                "sp500_cpi": "datahub.io Shiller S&P 500 (1871+) + Yahoo ^GSPC live",
                "commodities": "FRED PPIACO (1913+)", "deflator": "Shiller CPI + FRED CPIAUCNS",
                "silver_copper": "Yahoo SI=F / HG=F daily (2000+)",
            },
            "coverage_note": "S&P/Or = 1871+ · MP/Actions & MP réel = 1913+ · "
                             "Or nominal & Or/Argent & Cuivre/Or = daily 2000+ (futures Yahoo).",
            "palette": COL,
        },
        "composite": composite,
        "table": table,
        "series": {
            "sp_gold": to_series(sp_gold),
            "comm_eq": to_series(comm_eq),
            "comm_real": to_series(comm_real),
            "gold_nominal": gold_nom_series,
            "gold_silver": to_series_d(gold_silver),
            "copper_gold": to_series_d(copper_gold),
            "sp500": to_series(sp),
            "gold": to_series(gold_m),
        },
        "pendule_hist": pendule_hist,
        "seasonality": seasonality,
        "secular_cycles": secular,
    }
    return payload


# ─────────────────────── Merge / sanity / write ────────────────
def merge_preserve(new):
    if not CACHE_JSON.exists():
        return new
    try:
        old = json.loads(CACHE_JSON.read_text())
    except Exception:
        return new
    for k, sub in (("series", True),):
        for sk in (new.get("series") or {}):
            if not new["series"].get(sk) and old.get("series", {}).get(sk):
                new["series"][sk] = old["series"][sk]
                log(f"merge-preserve: series.{sk} conservé")
    # Saisonnalité : ne jamais régresser l'historique (ex. WorldBank indispo → argent/cuivre
    # retomberaient à 2000+). On ré-ajoute les années anciennes absentes du nouveau calcul.
    on, nn = old.get("seasonality") or {}, new.get("seasonality") or {}
    for asset, oyears in on.items():
        nyears = nn.setdefault(asset, {})
        added = 0
        for y, mrow in oyears.items():
            if y not in nyears:
                nyears[y] = mrow
                added += 1
        if added:
            log(f"merge-preserve: seasonality.{asset} +{added} années conservées")
    return new


def sanity(p):
    s = p["series"]
    assert len(s["sp_gold"]) > 500, "sp_gold trop court"
    assert len(s["gold_nominal"]) > 500, "gold trop court"
    sg = [r[1] for r in s["sp_gold"]]
    assert 0.05 < min(sg) and max(sg) < 20, f"sp_gold hors plage ({min(sg):.2f}..{max(sg):.2f})"
    assert p["composite"]["value"] is not None, "composite vide"
    assert any(r["value"] is not None for r in p["table"]), "table vide"


def write_outputs(p):
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    tmp_j = CACHE_DIR / ".tradfi_cycle.tmp.json"
    tmp_j.write_text(json.dumps(p, separators=(",", ":"), ensure_ascii=False))
    tmp_j.replace(CACHE_JSON)
    js = (f"/* tradfi_cycle_cache.js — {p['meta']['updated_at']} */\n"
          f"window.__TRADFI_CYCLE__={json.dumps(p, separators=(',', ':'), ensure_ascii=False)};\n")
    tmp_s = CACHE_DIR / ".tradfi_cycle.tmp.js"
    tmp_s.write_text(js)
    tmp_s.replace(CACHE_JS)
    if SITE_DIR.exists():
        for name in ("tradfi_cycle_cache.json", "tradfi_cycle_cache.js"):
            link, target = SITE_DIR / name, CACHE_DIR / name
            try:
                if link.is_symlink() or link.exists():
                    link.unlink()
                link.symlink_to(target)
            except OSError:
                import shutil
                shutil.copy2(target, link)


def main():
    if CACHE_JSON.exists() and CACHE_JS.exists() and "--force" not in sys.argv:
        age = (time.time() - CACHE_JSON.stat().st_mtime) / 3600
        if age < CACHE_MAX_HOURS:
            log(f"Cache frais ({age:.1f}h) — skip (--force pour forcer)")
            return
    t0 = time.time()
    try:
        payload = build()
    except Exception as e:
        log(f"FATAL build: {e}")
        sys.exit(2)
    payload = merge_preserve(payload)
    try:
        sanity(payload)
    except AssertionError as e:
        log(f"FATAL sanity: {e} — cache NON écrit")
        sys.exit(3)
    write_outputs(payload)
    c = payload["composite"]
    log(f"OK en {time.time()-t0:.1f}s · pendule={c['value']} ({c['label']})")
    for r in payload["table"]:
        log(f"   {r['label']:>32} = {str(r['display']):>10}  score {r['score']}  [{r['status']}]")


if __name__ == "__main__":
    main()
