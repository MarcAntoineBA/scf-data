#!/usr/bin/env python3
"""Cache « Adoption réelle de l'IA » pour le chapitre Thèse · La bulle IA (§ 11).

Mesure l'adoption ET l'intensité d'usage de l'IA, à plusieurs niveaux, avec des
sources TOUTES live + gratuites + auditables (chaque endpoint est ouvrable à la
main / curl, sans clé). Le but du chapitre : montrer que l'adoption est large
mais superficielle (chatbot) et profonde mais étroite (code agentique) — ce qui
matérialise le « mirage du ROI » du § 6.

Sources (toutes vérifiées 2026-06-17, valeurs de repli = dernières valeurs réelles) :

  CODE AGENTIQUE (la pointe de la pyramide d'usage)
   1. npm downloads API  → api.npmjs.org (no-auth, daily, historique 365 j)
        @anthropic-ai/claude-code, @openai/codex, @google/gemini-cli,
        @github/copilot (CLI terminal), opencode-ai, @qwen-code/qwen-code
   2. PyPI downloads     → pypistats.org  (aider-chat, open-interpreter)
   3. GitHub stars       → api.github.com (snapshot + diff historique)
        anthropics/claude-code, google-gemini/gemini-cli, openai/codex,
        cline/cline, block/goose, Aider-AI/aider, QwenLM/qwen-code

  INTENSITÉ D'USAGE (quel type d'usage consomme les tokens)
   4. OpenRouter rankings → openrouter.ai/api/frontend/v1/rankings/* (server-side)
        - use-case-category : mix programming vs roleplay vs marketing…
        - market-share      : part de tokens par fournisseur (DeepSeek/Anthropic…)
        - apps              : top apps (les agents de code dominent)

  ENTREPRISES
   5. US Census BTOS      → census.gov/hfp/btos (xlsx + API JSON)
        % de firmes US utilisant l'IA (bihebdomadaire, série depuis sept-2023)
        + adoption par fonction (AI Supplement) + par secteur

  ADMINISTRATIONS PUBLIQUES
   6. Inventaire IA fédéral US → raw.githubusercontent.com/ombegov (CSV)
        710 (2023) → 2 133 (2024), par agence / stade / impact
   7. UK ATRS            → gov.uk/api/search.json (compte de registres IA publics)
   8. USAspending        → api.usaspending.gov (contrats fédéraux IA, $/agence)

  MARCHÉ DU TRAVAIL
   9. Indeed Hiring Lab  → raw.githubusercontent.com/hiring-lab (part d'offres IA)

  CONSTANTES SOURCÉES (publications périodiques, valeurs vérifiées + lien source)
        Pew (34 % adultes US), WAU/MAU ChatGPT/Gemini/Claude, Anthropic Economic
        Index (split work/perso/études + collaboration), étude OpenAI NBER
        (codage 4,2 % sur ChatGPT vs 36 % sur Claude), McKinsey / Ramp / FactSet,
        Stack Overflow Dev Survey, Stanford HAI.

Sortie : ai_adoption_cache.json + .js (window.__AI_ADOPTION__), lus par
These_BulleIA.html. Lancé par scf.ai_adoption (toutes les ~6 h).
"""
import csv
import io
import json
import shutil
import sys
import time
from collections import OrderedDict, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import requests

# ── timeout global (sécurité launchd) ───────────────────────────────
import signal as _signal
def _timeout_handler(signum, frame):
    sys.stderr.write("[fatal] global timeout (25 min) — abort\n")
    sys.exit(2)
try:
    _signal.signal(_signal.SIGALRM, _timeout_handler)
    _signal.alarm(25 * 60)
except Exception:
    pass

CACHE_DIR = Path.home() / "Library" / "Caches" / "site_crypto_finance"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
OUT_JSON = CACHE_DIR / "ai_adoption_cache.json"
OUT_JS   = CACHE_DIR / "ai_adoption_cache.js"

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_0) SiteCryptoFinance-AIAdoption/1.0"
HEADERS = {"User-Agent": UA, "Accept": "application/json, text/csv, */*"}

# ════════════════════════════════════════════════════════════════════
#  CONFIG : ce qu'on suit
# ════════════════════════════════════════════════════════════════════

# (npm scope/name, libellé, éditeur)
NPM_TOOLS = [
    ("@openai/codex",            "Codex CLI",     "OpenAI"),
    ("@anthropic-ai/claude-code","Claude Code",   "Anthropic"),
    ("opencode-ai",              "opencode",      "SST (open-source)"),
    ("@github/copilot",          "Copilot CLI",   "GitHub"),
    ("@google/gemini-cli",       "Gemini CLI",    "Google"),
    ("@qwen-code/qwen-code",     "Qwen Code",     "Alibaba (open-source)"),
]

# (pkg PyPI, libellé)
PYPI_TOOLS = [
    ("aider-chat",       "aider"),
    ("open-interpreter", "Open Interpreter"),
]

# (repo, libellé)
GH_REPOS = [
    ("anthropics/claude-code",   "Claude Code"),
    ("google-gemini/gemini-cli", "Gemini CLI"),
    ("openai/codex",             "Codex"),
    ("cline/cline",              "Cline"),
    ("block/goose",              "goose"),
    ("Aider-AI/aider",           "aider"),
    ("QwenLM/qwen-code",         "Qwen Code"),
]

# Apps OpenRouter considérées « codage » (pour calculer la part codage des tokens)
CODING_APPS = {
    "claude code", "codex", "cline", "kilo code", "roo code", "opencode",
    "cursor", "aider", "continue", "windsurf", "kilocode", "roocode",
    "amp", "qwen code", "gemini cli", "copilot", "void", "zed",
}

OR_CATEGORIES = [
    "programming", "roleplay", "marketing", "technology",
    "science", "finance", "legal", "translation",
]

# ── Eurostat isoc_eb_ai : adoption de l'IA par les entreprises (10+), par pays ──
# Source officielle, auditable (endpoint ouvrable sans clé), annuelle (2021→).
# ATTENTION : la définition de la question IA a changé en 2023 → 2021 n'est PAS
# strictement comparable aux années suivantes (cf. "comparable_from": "2023").
EUROSTAT_BASE = ("https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/"
                 "data/isoc_eb_ai?format=JSON&lang=EN&size_emp=GE10&unit=PC_ENT")

GEO_FR = {
    "EU27_2020": "Union européenne (27)", "EA": "Zone euro", "BE": "Belgique",
    "BG": "Bulgarie", "CZ": "Tchéquie", "DK": "Danemark", "DE": "Allemagne",
    "EE": "Estonie", "IE": "Irlande", "EL": "Grèce", "ES": "Espagne", "FR": "France",
    "HR": "Croatie", "IT": "Italie", "CY": "Chypre", "LV": "Lettonie", "LT": "Lituanie",
    "LU": "Luxembourg", "HU": "Hongrie", "MT": "Malte", "NL": "Pays-Bas", "AT": "Autriche",
    "PL": "Pologne", "PT": "Portugal", "RO": "Roumanie", "SI": "Slovénie", "SK": "Slovaquie",
    "FI": "Finlande", "SE": "Suède", "NO": "Norvège", "IS": "Islande", "CH": "Suisse",
    "BA": "Bosnie-Herzégovine", "ME": "Monténégro", "AL": "Albanie", "RS": "Serbie",
    "MK": "Macédoine du Nord", "TR": "Türkiye", "XK": "Kosovo", "UK": "Royaume-Uni",
}
# Freins au non-usage (indic_is Eurostat → libellé FR)
BARRIERS_FR = OrderedDict([
    ("E_AI_BLE",  "Manque d'expertise"),
    ("E_AI_BLEG", "Flou juridique"),
    ("E_AI_BCDP", "Craintes vie privée / RGPD"),
    ("E_AI_BDDT", "Données insuffisantes"),
    ("E_AI_BINC", "Incompatibilité avec l'existant"),
    ("E_AI_BCST", "Coût trop élevé"),
    ("E_AI_BEC",  "Considérations éthiques"),
    ("E_AI_BNU",  "IA jugée inutile"),
])
# Finalités d'usage (indic_is Eurostat → libellé FR)
PURPOSES_FR = OrderedDict([
    ("E_AI_PMS",  "Marketing / ventes"),
    ("E_AI_PBAM", "Administration / gestion"),
    ("E_AI_PFIN", "Comptabilité / finance"),
    ("E_AI_PPP",  "Production"),
    ("E_AI_PITS", "Sécurité informatique"),
    ("E_AI_PRDI", "R&D / innovation"),
    ("E_AI_PLOG", "Logistique"),
    ("E_AI_PHR",  "RH / recrutement"),
])

# ════════════════════════════════════════════════════════════════════
#  VALEURS DE REPLI (dernières valeurs réelles vérifiées 2026-06-17)
#  Utilisées SEULEMENT si le fetch live échoue — sinon écrasées.
# ════════════════════════════════════════════════════════════════════

FALLBACK = {
    "agentic": {
        "npm": {
            "@openai/codex":            {"label": "Codex CLI",   "vendor": "OpenAI",            "month": 44852360, "history": []},
            "@anthropic-ai/claude-code":{"label": "Claude Code", "vendor": "Anthropic",         "month": 41018843, "history": []},
            "opencode-ai":              {"label": "opencode",    "vendor": "SST (open-source)", "month": 7267307,  "history": []},
            "@github/copilot":          {"label": "Copilot CLI", "vendor": "GitHub",            "month": 6676336,  "history": []},
            "@google/gemini-cli":       {"label": "Gemini CLI",  "vendor": "Google",            "month": 2858041,  "history": []},
            "@qwen-code/qwen-code":     {"label": "Qwen Code",   "vendor": "Alibaba (open-source)", "month": 432147, "history": []},
        },
        "pypi": {
            "aider-chat":       {"label": "aider",            "month": 890389, "week": 181690, "day": 30440},
            "open-interpreter": {"label": "Open Interpreter", "month": 205122, "week": 56081,  "day": 10445},
        },
        "github_stars": {
            "anthropics/claude-code":   {"label": "Claude Code", "stars": 133011, "forks": 21513},
            "google-gemini/gemini-cli": {"label": "Gemini CLI",  "stars": 105361, "forks": 14096},
            "openai/codex":             {"label": "Codex",       "stars": 91736,  "forks": 13556},
            "cline/cline":              {"label": "Cline",       "stars": 63451,  "forks": 6711},
            "block/goose":              {"label": "goose",       "stars": 49676,  "forks": 5260},
            "Aider-AI/aider":           {"label": "aider",       "stars": 46381,  "forks": 4611},
            "QwenLM/qwen-code":         {"label": "Qwen Code",   "stars": 25316,  "forks": 2525},
        },
        "stars_history": [],
        "no_public_metric": [
            {"name": "Cursor CLI", "vendor": "Anysphere", "why": "distribution via installeur curl (cursor.com/install) — aucun compteur public"},
            {"name": "Antigravity", "vendor": "Google", "why": "IDE binaire en preview — aucune métrique publiée ; embarque Gemini CLI (visible en partie via @google/gemini-cli)"},
        ],
    },
    "intensity": {
        "or_category_mix": [
            {"category": "programming", "tokens_b": 5960.3, "pct": 0.0},
            {"category": "roleplay",    "tokens_b": 0.0,    "pct": 0.0},
            {"category": "marketing",   "tokens_b": 0.0,    "pct": 0.0},
        ],
        "or_programming_share": {"weeks": [], "share_pct": [], "programming_b": []},
        "or_programming_models": [
            {"model": "xiaomi/mimo-v2.5",        "pct": 20.8, "tokens_b": 1239.5},
            {"model": "minimax/minimax-m3",      "pct": 14.9, "tokens_b": 890.8},
            {"model": "anthropic/claude-4.7-opus","pct": 7.0, "tokens_b": 418.2},
            {"model": "tencent/hy3-preview",     "pct": 6.7,  "tokens_b": 399.4},
            {"model": "deepseek/deepseek-v4-pro","pct": 6.6,  "tokens_b": 396.3},
            {"model": "Others",                  "pct": 28.2, "tokens_b": 1682.4},
        ],
        "or_market_share": {
            "weeks": [], "latest": [
                {"author": "deepseek", "pct": 18.0}, {"author": "anthropic", "pct": 15.9},
                {"author": "others", "pct": 12.8}, {"author": "minimax", "pct": 10.1},
                {"author": "xiaomi", "pct": 9.5}, {"author": "google", "pct": 9.3},
            ], "series": {},
        },
        "or_top_apps": [
            {"title": "Hermes Agent", "tokens_b": 6501.6, "requests": 89200000, "rank": 1, "coding": False},
            {"title": "Kilo Code",    "tokens_b": 2038.1, "requests": 35300000, "rank": 4, "coding": True},
            {"title": "OpenClaw",     "tokens_b": 1417.9, "requests": 26300000, "rank": 5, "coding": False},
            {"title": "Claude Code",  "tokens_b": 1052.9, "requests": 12600000, "rank": 6, "coding": True},
            {"title": "pi",           "tokens_b": 571.1,  "requests": 0,        "rank": 7, "coding": False},
            {"title": "Descript",     "tokens_b": 562.0,  "requests": 0,        "rank": 8, "coding": False},
            {"title": "Cline",        "tokens_b": 234.8,  "requests": 0,        "rank": 12, "coding": True},
            {"title": "Codex",        "tokens_b": 161.6,  "requests": 0,        "rank": 14, "coding": True},
        ],
        "or_coding_apps_share_pct": None,
    },
    "enterprise": {
        "btos": {
            "periods":     ["202604", "202605", "202606", "202607", "202608", "202609", "202610", "202611"],
            "labels":      ["fév-26", "fév-26", "mar-26", "mar-26", "avr-26", "avr-26", "mai-26", "mai-26"],
            "ai_now_yes":  [17.5, 18.2, 18.9, 19.1, 19.8, 19.8, 19.5, 20.6],
            "ai_6mo_yes":  [None, None, None, None, None, None, None, 23.7],
            "wording_break": "202524",
        },
        "btos_by_function": [
            {"function": "Vente & marketing",          "pct": 14.3},
            {"function": "Stratégie & dév. business",   "pct": 12.4},
            {"function": "IT",                          "pct": 11.4},
            {"function": "R&D",                         "pct": 11.2},
            {"function": "Service client",              "pct": 7.6},
            {"function": "Finance & comptabilité",      "pct": 6.6},
            {"function": "Fourniture de services",      "pct": 6.4},
            {"function": "RH",                          "pct": 4.2},
            {"function": "Production de biens",         "pct": 1.7},
        ],
        "btos_by_sector": [],
        "benchmarks": {
            "mckinsey_2025_pct": 88, "mckinsey_genai_pct": 72,
            "ramp_2026_pct": 54.2, "factset_q1_2026_calls": 337,
            "factset_q1_2026_total": 498, "factset_q1_2026_pct": 68,
        },
    },
    "public_admin": {
        "us_federal_inventory": {
            "y2023": {"total": 710},
            "y2024": {
                "total": 2133, "agencies": 41,
                "top_agencies": [
                    {"name": "HHS", "count": 271}, {"name": "DOJ", "count": 240},
                    {"name": "VA", "count": 229}, {"name": "DHS", "count": 183},
                    {"name": "Interior", "count": 180}, {"name": "USAID", "count": 137},
                    {"name": "Agriculture", "count": 89}, {"name": "Energy", "count": 79},
                    {"name": "Labor", "count": 70}, {"name": "Transportation", "count": 66},
                ],
                "by_stage": [
                    {"stage": "Exploitation & maintenance", "count": 874},
                    {"stage": "Acquisition / développement", "count": 404},
                    {"stage": "Initié", "count": 370},
                    {"stage": "Implémentation / évaluation", "count": 287},
                    {"stage": "Retiré", "count": 142},
                    {"stage": "Planifié", "count": 20},
                ],
                "by_impact": [
                    {"impact": "Ni l'un ni l'autre", "count": 1722},
                    {"impact": "Droits + sécurité", "count": 195},
                    {"impact": "Impactant les droits", "count": 131},
                    {"impact": "Impactant la sécurité", "count": 20},
                ],
            },
        },
        "uk_atrs": {"total": 133, "history": []},
        "usaspending_ai": {
            "fy": [2024], "count": [372], "total_usd_m": [],
            "by_agency_latest": [
                {"name": "Défense (DoD)", "usd_m": 187.5}, {"name": "Trésor", "usd_m": 21.8},
                {"name": "Santé (HHS)", "usd_m": 20.6}, {"name": "Commerce", "usd_m": 16.7},
                {"name": "Anciens combattants (VA)", "usd_m": 14.5}, {"name": "Sécurité int. (DHS)", "usd_m": 10.9},
            ],
            "latest_fy": 2024,
        },
        "context": {"stanford_federal_regs_2024": 59, "stanford_state_laws_2024": 131},
    },
    "personal": {
        "pew_chatgpt_ever_used": [
            {"wave": "juil-2023", "pct": 18}, {"wave": "fév-2024", "pct": 23}, {"wave": "mar-2025", "pct": 34},
        ],
        "assistant_users": [
            {"date": "2025-02", "product": "ChatGPT", "metric": "WAU", "value_m": 400, "source": "OpenAI"},
            {"date": "2025-10", "product": "ChatGPT", "metric": "WAU", "value_m": 800, "source": "OpenAI"},
            {"date": "2026-02", "product": "ChatGPT", "metric": "WAU", "value_m": 900, "source": "OpenAI (27 fév 2026)"},
            {"date": "2025-05", "product": "Gemini", "metric": "MAU", "value_m": 400, "source": "Google I/O"},
            {"date": "2026-02", "product": "Gemini", "metric": "MAU", "value_m": 750, "source": "Alphabet Q4-25"},
            {"date": "2025", "product": "Claude", "metric": "MAU (est.)", "value_m": 20, "source": "Sacra / Business of Apps"},
        ],
        "aei_use_case": [
            {"use": "Travail", "pct": 45.2}, {"use": "Personnel", "pct": 42.3}, {"use": "Études", "pct": 12.4},
        ],
        "aei_collaboration": [
            {"mode": "Directif (délégation)", "pct": 32.6}, {"mode": "Itération sur tâche", "pct": 25.6},
            {"mode": "Apprentissage", "pct": 22.4}, {"mode": "Boucle de feedback", "pct": 11.5},
            {"mode": "Validation", "pct": 4.9},
        ],
        "coding_contrast": {"chatgpt_pct": 4.2, "claude_ai_pct": 36, "claude_api_pct": 44},
        "chatgpt_message_types": [
            {"type": "Conseil pratique", "pct": 29}, {"type": "Rédaction", "pct": 24},
            {"type": "Recherche d'info", "pct": 24}, {"type": "Programmation", "pct": 4.2},
        ],
    },
    "labor": {
        "indeed_ai_share": {
            "dates": ["2019-01", "2022-01", "2024-01", "2025-01", "2026-01", "2026-05"],
            "series": {
                "US": [1.72, 2.50, 1.93, 2.70, 4.41, 5.68],
                "GB": [None, None, None, None, None, 9.20],
                "CA": [None, None, None, None, None, 16.97],
            },
            "latest_date": "2026-05-31",
            "latest_by_country": [
                {"country": "Canada", "pct": 16.97}, {"country": "Irlande", "pct": 13.92},
                {"country": "Royaume-Uni", "pct": 9.20}, {"country": "Australie", "pct": 7.23},
                {"country": "États-Unis", "pct": 5.68}, {"country": "Allemagne", "pct": 4.98},
                {"country": "Italie", "pct": 4.44}, {"country": "France", "pct": 4.18},
                {"country": "Pays-Bas", "pct": 3.08},
            ],
        },
        "so_survey": {"ai_use_2024": 76, "ai_use_2025": 84, "daily_pct": 51},
    },
}

# Repli « Géographie » (Eurostat isoc_eb_ai) — dernier instantané vérifié 2026-07-04.
# JSON embarqué (évite les soucis true/false littéraux) ; écrasé par le fetch live.
_GEO_FALLBACK_JSON = r'''{"eurostat": {"years": ["2021", "2023", "2024", "2025"], "comparable_from": "2023", "latest_year": "2025", "adoption": [{"code": "DK", "name": "Danemark", "vals": {"2021": 23.89, "2023": 15.17, "2024": 27.58, "2025": 42.03}}, {"code": "FI", "name": "Finlande", "vals": {"2021": 15.79, "2023": 15.1, "2024": 24.37, "2025": 37.82}}, {"code": "SE", "name": "Suède", "vals": {"2021": 9.92, "2023": 10.37, "2024": 25.09, "2025": 35.04}}, {"code": "BE", "name": "Belgique", "vals": {"2021": 10.32, "2023": 13.81, "2024": 24.71, "2025": 34.54}}, {"code": "LU", "name": "Luxembourg", "vals": {"2021": 13.0, "2023": 14.45, "2024": 23.73, "2025": 33.61}}, {"code": "NL", "name": "Pays-Bas", "vals": {"2021": 13.1, "2023": 14.1, "2024": 23.06, "2025": 33.21}}, {"code": "AT", "name": "Autriche", "vals": {"2021": 8.83, "2023": 10.79, "2024": 20.27, "2025": 29.95}}, {"code": "NO", "name": "Norvège", "vals": {"2021": 10.82, "2023": 9.17, "2024": 20.77, "2025": 28.89}}, {"code": "DE", "name": "Allemagne", "vals": {"2021": 10.56, "2023": 11.55, "2024": 19.75, "2025": 25.97}}, {"code": "EE", "name": "Estonie", "vals": {"2021": 2.77, "2023": 5.19, "2024": 13.89, "2025": 23.4}}, {"code": "SI", "name": "Slovénie", "vals": {"2021": 11.73, "2023": 11.37, "2024": 20.89, "2025": 21.61}}, {"code": "MT", "name": "Malte", "vals": {"2021": 10.16, "2023": 13.17, "2024": 17.3, "2025": 21.51}}, {"code": "LT", "name": "Lituanie", "vals": {"2021": 4.45, "2023": 4.86, "2024": 8.76, "2025": 21.3}}, {"code": "ES", "name": "Espagne", "vals": {"2021": 7.67, "2023": 9.18, "2024": 11.31, "2025": 20.27}}, {"code": "IE", "name": "Irlande", "vals": {"2021": 7.88, "2023": 8.01, "2024": 14.9, "2025": 19.64}}, {"code": "FR", "name": "France", "vals": {"2021": 6.67, "2023": 5.88, "2024": 9.91, "2025": 18.16}}, {"code": "SK", "name": "Slovaquie", "vals": {"2021": 5.19, "2023": 7.04, "2024": 10.78, "2025": 18.0}}, {"code": "CZ", "name": "Tchéquie", "vals": {"2021": 4.46, "2023": 5.9, "2024": 11.26, "2025": 17.6}}, {"code": "IT", "name": "Italie", "vals": {"2021": 6.17, "2023": 5.05, "2024": 8.2, "2025": 16.4}}, {"code": "HR", "name": "Croatie", "vals": {"2021": 8.74, "2023": 7.89, "2024": 11.76, "2025": 15.19}}, {"code": "LV", "name": "Lettonie", "vals": {"2021": 3.72, "2023": 4.53, "2024": 8.83, "2025": 12.21}}, {"code": "PT", "name": "Portugal", "vals": {"2021": 7.2, "2023": 7.86, "2024": 8.63, "2025": 11.54}}, {"code": "BA", "name": "Bosnie-Herzégovine", "vals": {"2021": 2.07, "2023": 5.34, "2024": 6.36, "2025": 10.78}}, {"code": "HU", "name": "Hongrie", "vals": {"2021": 2.98, "2023": 3.68, "2024": 7.41, "2025": 10.37}}, {"code": "RS", "name": "Serbie", "vals": {"2021": 0.9, "2023": 1.82, "2024": 6.95, "2025": 10.12}}, {"code": "ME", "name": "Monténégro", "vals": {"2021": 3.34, "2023": 5.61, "2024": 7.91, "2025": 10.05}}, {"code": "CY", "name": "Chypre", "vals": {"2021": 2.59, "2023": 4.67, "2024": 7.9, "2025": 9.27}}, {"code": "AL", "name": "Albanie", "vals": {"2021": 3.67, "2024": 8.91, "2025": 8.99}}, {"code": "EL", "name": "Grèce", "vals": {"2021": 2.61, "2023": 3.98, "2024": 9.81, "2025": 8.93}}, {"code": "BG", "name": "Bulgarie", "vals": {"2021": 3.29, "2023": 3.62, "2024": 6.47, "2025": 8.55}}, {"code": "PL", "name": "Pologne", "vals": {"2021": 2.86, "2023": 3.67, "2024": 5.9, "2025": 8.36}}, {"code": "TR", "name": "Türkiye", "vals": {"2021": 2.69, "2023": 5.51, "2024": 4.42, "2025": 7.41}}, {"code": "RO", "name": "Roumanie", "vals": {"2021": 1.38, "2023": 1.51, "2024": 3.07, "2025": 5.21}}], "aggregates": {"EU27_2020": {"code": "EU27_2020", "name": "Union européenne (27)", "vals": {"2021": 7.65, "2023": 8.06, "2024": 13.48, "2025": 19.95}}, "EA": {"code": "EA", "name": "Zone euro", "vals": {"2021": 8.36, "2023": 8.84, "2024": 14.39, "2025": 21.37}}}, "generative": [{"code": "FI", "name": "Finlande", "pct": 21.55, "agg": false}, {"code": "SE", "name": "Suède", "pct": 18.67, "agg": false}, {"code": "DK", "name": "Danemark", "pct": 17.66, "agg": false}, {"code": "BE", "name": "Belgique", "pct": 17.32, "agg": false}, {"code": "NL", "name": "Pays-Bas", "pct": 14.92, "agg": false}, {"code": "LU", "name": "Luxembourg", "pct": 14.21, "agg": false}, {"code": "DE", "name": "Allemagne", "pct": 13.52, "agg": false}, {"code": "AT", "name": "Autriche", "pct": 13.08, "agg": false}, {"code": "NO", "name": "Norvège", "pct": 13.05, "agg": false}, {"code": "LT", "name": "Lituanie", "pct": 11.49, "agg": false}, {"code": "ES", "name": "Espagne", "pct": 11.38, "agg": false}, {"code": "SK", "name": "Slovaquie", "pct": 10.93, "agg": false}, {"code": "EA", "name": "Zone euro", "pct": 10.25, "agg": true}, {"code": "CZ", "name": "Tchéquie", "pct": 9.68, "agg": false}, {"code": "EU27_2020", "name": "Union européenne (27)", "pct": 9.55, "agg": true}, {"code": "FR", "name": "France", "pct": 7.5, "agg": false}, {"code": "MT", "name": "Malte", "pct": 7.13, "agg": false}, {"code": "HR", "name": "Croatie", "pct": 7.04, "agg": false}, {"code": "IT", "name": "Italie", "pct": 7.04, "agg": false}, {"code": "EE", "name": "Estonie", "pct": 6.46, "agg": false}, {"code": "SI", "name": "Slovénie", "pct": 6.05, "agg": false}, {"code": "PT", "name": "Portugal", "pct": 5.88, "agg": false}, {"code": "IE", "name": "Irlande", "pct": 5.65, "agg": false}, {"code": "LV", "name": "Lettonie", "pct": 5.48, "agg": false}, {"code": "BA", "name": "Bosnie-Herzégovine", "pct": 5.01, "agg": false}, {"code": "TR", "name": "Türkiye", "pct": 4.75, "agg": false}, {"code": "CY", "name": "Chypre", "pct": 4.68, "agg": false}, {"code": "HU", "name": "Hongrie", "pct": 4.65, "agg": false}, {"code": "BG", "name": "Bulgarie", "pct": 4.52, "agg": false}, {"code": "EL", "name": "Grèce", "pct": 4.02, "agg": false}, {"code": "RS", "name": "Serbie", "pct": 3.95, "agg": false}, {"code": "AL", "name": "Albanie", "pct": 3.53, "agg": false}, {"code": "RO", "name": "Roumanie", "pct": 3.03, "agg": false}, {"code": "PL", "name": "Pologne", "pct": 2.75, "agg": false}, {"code": "ME", "name": "Monténégro", "pct": 1.85, "agg": false}], "barriers_eu": [{"code": "E_AI_BLE", "label": "Manque d'expertise", "pct": 7.76}, {"code": "E_AI_BLEG", "label": "Flou juridique", "pct": 5.92}, {"code": "E_AI_BCDP", "label": "Craintes vie privée / RGPD", "pct": 5.82}, {"code": "E_AI_BDDT", "label": "Données insuffisantes", "pct": 4.8}, {"code": "E_AI_BINC", "label": "Incompatibilité avec l'existant", "pct": 4.59}, {"code": "E_AI_BCST", "label": "Coût trop élevé", "pct": 4.24}, {"code": "E_AI_BEC", "label": "Considérations éthiques", "pct": 2.72}, {"code": "E_AI_BNU", "label": "IA jugée inutile", "pct": 1.96}], "purposes_eu": [{"code": "E_AI_PMS", "label": "Marketing / ventes", "pct": 6.92}, {"code": "E_AI_PBAM", "label": "Administration / gestion", "pct": 6.19}, {"code": "E_AI_PFIN", "label": "Comptabilité / finance", "pct": 4.61}, {"code": "E_AI_PPP", "label": "Production", "pct": 4.14}, {"code": "E_AI_PITS", "label": "Sécurité informatique", "pct": 3.94}, {"code": "E_AI_PRDI", "label": "R&D / innovation", "pct": 3.8}, {"code": "E_AI_PLOG", "label": "Logistique", "pct": 1.21}], "unit": "PC_ENT", "size": "GE10", "src": "isoc_eb_ai"}}'''
FALLBACK["geo"] = json.loads(_GEO_FALLBACK_JSON)

# ════════════════════════════════════════════════════════════════════
#  Helpers HTTP
# ════════════════════════════════════════════════════════════════════

def get_json(url, timeout=25, params=None):
    r = requests.get(url, headers=HEADERS, params=params, timeout=timeout)
    r.raise_for_status()
    return r.json()

def get_text(url, timeout=40):
    r = requests.get(url, headers=HEADERS, timeout=timeout)
    r.raise_for_status()
    return r.text

def get_bytes(url, timeout=60):
    r = requests.get(url, headers=HEADERS, timeout=timeout)
    r.raise_for_status()
    return r.content

def post_json(url, body, timeout=40):
    r = requests.post(url, headers={"User-Agent": UA, "Content-Type": "application/json"},
                      data=json.dumps(body), timeout=timeout)
    r.raise_for_status()
    return r.json()


# ════════════════════════════════════════════════════════════════════
#  1. npm downloads (monthly buckets sur 1 an + headline last-month)
# ════════════════════════════════════════════════════════════════════

def fetch_npm(ok, failed):
    out = {}
    for scope_name, label, vendor in NPM_TOOLS:
        try:
            # headline : dernier mois glissant
            pt = get_json(f"https://api.npmjs.org/downloads/point/last-month/{scope_name}")
            month = int(pt.get("downloads", 0))
            # historique : 365 j -> buckets mensuels
            rng = get_json(f"https://api.npmjs.org/downloads/range/last-year/{scope_name}")
            buckets = OrderedDict()
            for d in rng.get("downloads", []):
                ym = d["day"][:7]
                buckets[ym] = buckets.get(ym, 0) + int(d.get("downloads", 0))
            history = [{"month": k, "downloads": v} for k, v in buckets.items()]
            out[scope_name] = {"label": label, "vendor": vendor, "month": month, "history": history}
            ok.append(f"npm:{scope_name}")
        except Exception as e:
            sys.stderr.write(f"[npm {scope_name}] {e}\n")
            failed.append(f"npm:{scope_name}")
    return out or None


# ════════════════════════════════════════════════════════════════════
#  2. PyPI downloads (pypistats — espacer les appels, rate-limit agressif)
# ════════════════════════════════════════════════════════════════════

def fetch_pypi(ok, failed):
    out = {}
    for i, (pkg, label) in enumerate(PYPI_TOOLS):
        try:
            if i:
                time.sleep(8)  # espacement anti-429
            j = get_json(f"https://pypistats.org/api/packages/{pkg}/recent", timeout=30)
            data = j.get("data", {})
            out[pkg] = {
                "label": label,
                "month": int(data.get("last_month", 0)),
                "week": int(data.get("last_week", 0)),
                "day": int(data.get("last_day", 0)),
            }
            ok.append(f"pypi:{pkg}")
        except Exception as e:
            sys.stderr.write(f"[pypi {pkg}] {e}\n")
            failed.append(f"pypi:{pkg}")
    return out or None


# ════════════════════════════════════════════════════════════════════
#  3. GitHub stars (snapshot ; l'historique est construit par diff)
# ════════════════════════════════════════════════════════════════════

def fetch_github(ok, failed):
    out = {}
    gh_headers = {"User-Agent": UA, "Accept": "application/vnd.github+json"}
    for repo, label in GH_REPOS:
        try:
            r = requests.get(f"https://api.github.com/repos/{repo}", headers=gh_headers,
                             timeout=20, allow_redirects=True)
            r.raise_for_status()
            j = r.json()
            out[repo] = {"label": label,
                         "stars": int(j.get("stargazers_count", 0)),
                         "forks": int(j.get("forks_count", 0))}
            ok.append(f"gh:{repo}")
        except Exception as e:
            sys.stderr.write(f"[gh {repo}] {e}\n")
            failed.append(f"gh:{repo}")
    return out or None


# ════════════════════════════════════════════════════════════════════
#  4. OpenRouter rankings (server-side ; CORS verrouillé navigateur)
# ════════════════════════════════════════════════════════════════════

OR_BASE = "https://openrouter.ai/api/frontend/v1/rankings"  # v1 depuis ~07/2026 (ancien /rankings → 404)

def fetch_openrouter(ok, failed):
    out = {}
    # 4a. mix par catégorie d'usage + part programming dans le temps
    try:
        cat_weeks = {}           # category -> {week: total_b}
        for cat in OR_CATEGORIES:
            j = get_json(f"{OR_BASE}/use-case-category", params={"category": cat}, timeout=30)
            wk = {}
            for pt in j.get("data", []):
                x = pt.get("x")
                ys = pt.get("ys", {})
                tot = sum(float(v) for v in ys.values()) / 1e9  # tokens -> milliards
                wk[x] = tot
            cat_weeks[cat] = wk
            time.sleep(0.5)
        # dernier point commun
        all_weeks = sorted(set().union(*[set(v.keys()) for v in cat_weeks.values()]))
        if all_weeks:
            last = all_weeks[-1]
            mix = [{"category": c, "tokens_b": round(cat_weeks[c].get(last, 0.0), 1)}
                   for c in OR_CATEGORIES]
            tot_last = sum(m["tokens_b"] for m in mix) or 1.0
            for m in mix:
                m["pct"] = round(100 * m["tokens_b"] / tot_last, 1)
            mix.sort(key=lambda m: -m["tokens_b"])
            out["or_category_mix"] = mix
            # part programming sur tout le trafic catégorisé, semaine par semaine
            weeks, share, prog_b = [], [], []
            for w in all_weeks:
                tot = sum(cat_weeks[c].get(w, 0.0) for c in OR_CATEGORIES)
                p = cat_weeks["programming"].get(w, 0.0)
                if tot > 0:
                    weeks.append(w); share.append(round(100 * p / tot, 1)); prog_b.append(round(p, 1))
            out["or_programming_share"] = {"weeks": weeks, "share_pct": share, "programming_b": prog_b}
        # top modèles dans programming (dernière semaine)
        jp = get_json(f"{OR_BASE}/use-case-category", params={"category": "programming"}, timeout=30)
        if jp.get("data"):
            ys = jp["data"][-1].get("ys", {})
            tot = sum(float(v) for v in ys.values()) or 1.0
            models = [{"model": k, "tokens_b": round(float(v) / 1e9, 1),
                       "pct": round(100 * float(v) / tot, 1)} for k, v in ys.items()]
            models.sort(key=lambda m: -m["tokens_b"])
            out["or_programming_models"] = models[:8]
        ok.append("openrouter:use-case-category")
    except Exception as e:
        sys.stderr.write(f"[openrouter category] {e}\n")
        failed.append("openrouter:use-case-category")

    # 4b. part de marché par fournisseur (séries hebdo)
    try:
        j = get_json(f"{OR_BASE}/market-share", timeout=30)
        data = j.get("data", [])
        weeks = [pt.get("x") for pt in data]
        series = defaultdict(list)
        authors = set()
        for pt in data:
            authors |= set(pt.get("ys", {}).keys())
        for pt in data:
            ys = pt.get("ys", {})
            for a in authors:
                series[a].append(round(float(ys.get(a, 0)) / 1e9, 1))
        latest = []
        if data:
            ys = data[-1].get("ys", {})
            tot = sum(float(v) for v in ys.values()) or 1.0
            latest = [{"author": a, "pct": round(100 * float(v) / tot, 1)} for a, v in ys.items()]
            latest.sort(key=lambda x: -x["pct"])
        out["or_market_share"] = {"weeks": weeks, "series": dict(series), "latest": latest[:8]}
        ok.append("openrouter:market-share")
    except Exception as e:
        sys.stderr.write(f"[openrouter market-share] {e}\n")
        failed.append("openrouter:market-share")

    # 4c. top apps (les agents de code dominent) + part codage
    try:
        j = get_json(f"{OR_BASE}/apps", timeout=30)
        week = j.get("data", {}).get("week", [])
        apps, tot_tok, cod_tok = [], 0.0, 0.0
        for e in week:
            tok_b = float(e.get("total_tokens", 0)) / 1e9
            title = (e.get("app", {}) or {}).get("title") or e.get("title") or "—"
            is_cod = title.strip().lower() in CODING_APPS
            tot_tok += tok_b
            if is_cod:
                cod_tok += tok_b
            apps.append({"title": title, "tokens_b": round(tok_b, 1),
                         "requests": int(e.get("total_requests", 0) or 0),
                         "rank": int(e.get("rank", 0) or 0), "coding": is_cod})
        apps.sort(key=lambda a: -a["tokens_b"])
        out["or_top_apps"] = apps[:15]
        out["or_coding_apps_share_pct"] = round(100 * cod_tok / tot_tok, 1) if tot_tok else None
        ok.append("openrouter:apps")
    except Exception as e:
        sys.stderr.write(f"[openrouter apps] {e}\n")
        failed.append("openrouter:apps")

    return out or None


# ════════════════════════════════════════════════════════════════════
#  5. US Census BTOS (National.xlsx — % firmes utilisant l'IA)
# ════════════════════════════════════════════════════════════════════

def _period_label(code):
    """202611 -> 'mai-26' (approx via quinzaine)."""
    try:
        y = int(code[:4]); n = int(code[4:])
        month = min(12, max(1, (n + 1) // 2))
        mois = ["jan", "fév", "mar", "avr", "mai", "juin",
                "juil", "août", "sep", "oct", "nov", "déc"][month - 1]
        return f"{mois}-{str(y)[2:]}"
    except Exception:
        return code

def fetch_btos(ok, failed):
    try:
        from openpyxl import load_workbook
    except Exception as e:
        sys.stderr.write(f"[btos] openpyxl absent: {e}\n")
        failed.append("btos:National.xlsx")
        return None
    try:
        raw = get_bytes("https://www.census.gov/hfp/btos/downloads/National.xlsx", timeout=40)
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb["Response Estimates"]
        rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
        wb.close()
        # repérer la ligne d'en-tête (cellules = codes période YYYYNN)
        hdr_idx, period_cols = None, []
        for i, row in enumerate(rows[:8]):
            codes = [(j, str(v)) for j, v in enumerate(row)
                     if v is not None and str(v).strip().isdigit() and len(str(v).strip()) == 6]
            if len(codes) >= 5:
                hdr_idx = i
                period_cols = codes  # liste (col_idx, code)
                break
        if hdr_idx is None:
            raise RuntimeError("en-tête période introuvable")
        # colonnes question/réponse : chercher dans les premières colonnes
        def cell(row, idx):
            return "" if idx >= len(row) or row[idx] is None else str(row[idx]).strip()
        # détecter idx question + réponse depuis l'en-tête
        header = rows[hdr_idx]
        q_idx = a_idx = None
        for j, v in enumerate(header):
            s = (str(v).strip().lower() if v is not None else "")
            if s == "question":
                q_idx = j
            if s == "answer":
                a_idx = j
        if q_idx is None:
            q_idx = 1
        if a_idx is None:
            a_idx = 3
        # codes période -> ordre chronologique (les colonnes sont newest-first)
        period_cols_sorted = sorted(period_cols, key=lambda c: c[1])  # ancien -> récent
        periods = [c[1] for c in period_cols_sorted]
        now_yes = {p: None for p in periods}
        six_yes = {p: None for p in periods}
        for row in rows[hdr_idx + 1:]:
            q = cell(row, q_idx).lower()
            a = cell(row, a_idx).lower()
            if "artificial intelligence" not in q or a != "yes":
                continue
            is_future = ("six months" in q) or ("next six" in q)
            target = six_yes if is_future else now_yes
            for col_idx, code in period_cols_sorted:
                val = cell(row, col_idx).replace("%", "").replace(",", ".")
                try:
                    target[code] = round(float(val), 1)
                except Exception:
                    pass
        # garder uniquement les périodes où ai_now existe
        keep = [p for p in periods if now_yes[p] is not None]
        if len(keep) < 4:
            raise RuntimeError("trop peu de points BTOS valides")
        out = {
            "periods": keep,
            "labels": [_period_label(p) for p in keep],
            "ai_now_yes": [now_yes[p] for p in keep],
            "ai_6mo_yes": [six_yes[p] for p in keep],
            "wording_break": "202524",
        }
        ok.append("btos:National.xlsx")
        return out
    except Exception as e:
        sys.stderr.write(f"[btos] {e}\n")
        failed.append("btos:National.xlsx")
        return None


# ════════════════════════════════════════════════════════════════════
#  6. Inventaire IA fédéral US (CSV GitHub OMB)
# ════════════════════════════════════════════════════════════════════

def _read_csv_rows(url, encoding):
    txt = get_bytes(url, timeout=60).decode(encoding, errors="replace")
    return list(csv.reader(io.StringIO(txt)))

def fetch_federal_inventory(ok, failed):
    out = {}
    base = "https://raw.githubusercontent.com/ombegov/2024-Federal-AI-Use-Case-Inventory/main/data"
    # 2023 baseline
    try:
        rows = _read_csv_rows(f"{base}/2023_consolidated_ai_inventory_raw.csv", "utf-8-sig")
        out["y2023"] = {"total": max(0, len(rows) - 1)}
        ok.append("fedinv:2023")
    except Exception as e:
        sys.stderr.write(f"[fedinv 2023] {e}\n")
        failed.append("fedinv:2023")
    # 2024 détaillé
    try:
        rows = _read_csv_rows(f"{base}/2024_consolidated_ai_inventory_raw_v2.csv", "cp1252")
        header = rows[0]
        idx = {h.strip().lower(): i for i, h in enumerate(header)}
        def col(*names):
            for n in names:
                for h, i in idx.items():
                    if n in h:
                        return i
            return None
        a_i = col("agency")  # 'Agency' (department-level en 2024)
        s_i = col("stage")
        # impact : colonnes 'rights'/'safety' ou une colonne unique
        data = rows[1:]
        agencies = defaultdict(int); stages = defaultdict(int)
        for r in data:
            if a_i is not None and a_i < len(r) and r[a_i].strip():
                agencies[r[a_i].strip()] += 1
            if s_i is not None and s_i < len(r) and r[s_i].strip():
                stages[r[s_i].strip()] += 1
        top = sorted(agencies.items(), key=lambda x: -x[1])[:10]
        y24 = {
            "total": len(data),
            "agencies": len(agencies),
            "top_agencies": [{"name": k, "count": v} for k, v in top],
        }
        if stages:
            y24["by_stage"] = [{"stage": k, "count": v}
                               for k, v in sorted(stages.items(), key=lambda x: -x[1])[:6]]
        out["y2024"] = y24
        ok.append("fedinv:2024")
    except Exception as e:
        sys.stderr.write(f"[fedinv 2024] {e}\n")
        failed.append("fedinv:2024")
    return out or None


# ════════════════════════════════════════════════════════════════════
#  7. UK ATRS (GOV.UK Search API)
# ════════════════════════════════════════════════════════════════════

def fetch_uk_atrs(ok, failed):
    try:
        j = get_json("https://www.gov.uk/api/search.json",
                     params={"filter_content_store_document_type": "algorithmic_transparency_record",
                             "count": "0"}, timeout=20)
        total = int(j.get("total", 0))
        if total <= 0:
            raise RuntimeError("total ATRS = 0")
        ok.append("uk:atrs")
        return {"total": total}
    except Exception as e:
        sys.stderr.write(f"[uk atrs] {e}\n")
        failed.append("uk:atrs")
        return None


# ════════════════════════════════════════════════════════════════════
#  8. USAspending (contrats fédéraux IA)
# ════════════════════════════════════════════════════════════════════

def _usa_count(fy):
    body = {
        "filters": {
            "keywords": ["artificial intelligence"],
            "award_type_codes": ["A", "B", "C", "D"],
            "time_period": [{"start_date": f"{fy-1}-10-01", "end_date": f"{fy}-09-30"}],
        }
    }
    j = post_json("https://api.usaspending.gov/api/v2/search/spending_by_award_count/", body, timeout=40)
    return int(j.get("results", {}).get("contracts", 0))

def fetch_usaspending(ok, failed):
    out = {"fy": [], "count": [], "by_agency_latest": [], "latest_fy": None}
    try:
        for fy in range(2020, 2026):
            try:
                c = _usa_count(fy)
                out["fy"].append(fy); out["count"].append(c)
                time.sleep(0.5)
            except Exception as e:
                sys.stderr.write(f"[usa fy{fy}] {e}\n")
        if not out["fy"]:
            raise RuntimeError("aucune année USAspending")
        out["latest_fy"] = out["fy"][-1]
        # $ par agence sur la dernière FY
        try:
            fy = out["latest_fy"]
            body = {
                "category": "awarding_agency",
                "filters": {
                    "keywords": ["artificial intelligence"],
                    "award_type_codes": ["A", "B", "C", "D"],
                    "time_period": [{"start_date": f"{fy-1}-10-01", "end_date": f"{fy}-09-30"}],
                },
                "limit": 8,
            }
            j = post_json("https://api.usaspending.gov/api/v2/search/spending_by_category/awarding_agency/",
                          body, timeout=40)
            agz = [{"name": r.get("name", "—"),
                    "usd_m": round(float(r.get("amount", 0)) / 1e6, 1)}
                   for r in j.get("results", []) if r.get("name")]
            if agz:
                out["by_agency_latest"] = agz
            else:
                out["by_agency_latest"] = FALLBACK["public_admin"]["usaspending_ai"]["by_agency_latest"]
        except Exception as e:
            sys.stderr.write(f"[usa agencies] {e}\n")
            out["by_agency_latest"] = FALLBACK["public_admin"]["usaspending_ai"]["by_agency_latest"]
        ok.append("usaspending:ai")
        return out
    except Exception as e:
        sys.stderr.write(f"[usaspending] {e}\n")
        failed.append("usaspending:ai")
        return None


# ════════════════════════════════════════════════════════════════════
#  9. Indeed Hiring Lab — part d'offres mentionnant l'IA
# ════════════════════════════════════════════════════════════════════

COUNTRY_FR = {"US": "États-Unis", "GB": "Royaume-Uni", "CA": "Canada", "FR": "France",
              "DE": "Allemagne", "AU": "Australie", "IE": "Irlande", "IT": "Italie", "NL": "Pays-Bas"}

def fetch_indeed(ok, failed):
    try:
        txt = get_text("https://raw.githubusercontent.com/hiring-lab/ai-tracker/main/AI_posting.csv", timeout=40)
        rdr = csv.DictReader(io.StringIO(txt))
        # détecter noms de colonnes
        fields = rdr.fieldnames or []
        date_c = next((f for f in fields if f.lower() == "date"), fields[0])
        ctry_c = next((f for f in fields if "country" in f.lower()), None)
        val_c = next((f for f in fields if "share" in f.lower() or "ai" in f.lower() and "share" in f.lower()), None)
        if val_c is None:
            val_c = next((f for f in fields if f not in (date_c, ctry_c)), None)
        # garder un point mensuel par pays (premier obs du mois)
        keep_countries = ["US", "GB", "CA", "FR", "DE"]
        monthly = defaultdict(dict)   # ym -> {country: val}
        seen = set()
        for row in rdr:
            c = (row.get(ctry_c) or "").strip().upper()
            if c not in keep_countries:
                continue
            d = (row.get(date_c) or "").strip()
            ym = d[:7]
            key = (c, ym)
            if key in seen:
                continue
            seen.add(key)
            try:
                monthly[ym][c] = round(float(row.get(val_c)), 2)
            except Exception:
                pass
        dates = sorted(monthly.keys())
        series = {c: [monthly[ym].get(c) for ym in dates] for c in keep_countries}
        # dernière photo cross-pays (toutes nations dispo, dernière date)
        rdr2 = csv.DictReader(io.StringIO(txt))
        last_date = ""
        last_vals = {}
        for row in rdr2:
            d = (row.get(date_c) or "").strip()
            c = (row.get(ctry_c) or "").strip().upper()
            try:
                v = round(float(row.get(val_c)), 2)
            except Exception:
                continue
            if d > last_date:
                last_date = d; last_vals = {}
            if d == last_date:
                last_vals[c] = v
        latest = sorted([{"country": COUNTRY_FR.get(k, k), "pct": v} for k, v in last_vals.items()],
                        key=lambda x: -x["pct"])
        out = {"dates": dates, "series": series, "latest_date": last_date, "latest_by_country": latest}
        ok.append("indeed:ai_posting")
        return out
    except Exception as e:
        sys.stderr.write(f"[indeed] {e}\n")
        failed.append("indeed:ai_posting")
        return None


# ════════════════════════════════════════════════════════════════════
#  10. Eurostat — adoption de l'IA par les entreprises, PAR PAYS
#      (isoc_eb_ai · % d'entreprises 10+ · JSON-stat · sans clé)
# ════════════════════════════════════════════════════════════════════

def _eurostat_val(d, **coords):
    """Valeur JSON-stat par coordonnées (dims non fournies = catégorie unique, idx 0)."""
    dims = d["dimension"]; ids = d["id"]
    sizes = [len(dims[i]["category"]["index"]) for i in ids]
    strides = [1] * len(ids)
    for i in range(len(ids) - 2, -1, -1):
        strides[i] = strides[i + 1] * sizes[i + 1]
    k = 0
    for pos, dim in enumerate(ids):
        cats = dims[dim]["category"]["index"]
        idx = cats.get(coords[dim]) if dim in coords else 0
        if idx is None:
            return None
        k += idx * strides[pos]
    return d["value"].get(str(k))

def fetch_eurostat_geo(ok, failed):
    out = {
        "years": [], "comparable_from": "2023", "latest_year": None,
        "adoption": [], "aggregates": {}, "generative": [],
        "barriers_eu": [], "purposes_eu": [],
        "unit": "PC_ENT", "size": "GE10", "src": "isoc_eb_ai",
    }
    got = 0
    # A) adoption « au moins 1 techno IA » (E_AI_TANY) : geo × time
    try:
        d = get_json(EUROSTAT_BASE + "&indic_is=E_AI_TANY", timeout=45)
        dims = d["dimension"]; gl = dims["geo"]["category"]["label"]
        gidx = dims["geo"]["category"]["index"]; tidx = dims["time"]["category"]["index"]
        years = sorted(tidx, key=lambda t: tidx[t])
        out["years"] = years; out["latest_year"] = years[-1]
        ly = years[-1]; adoption = []; aggs = {}
        for g in gidx:
            vv = {}
            for y in years:
                v = _eurostat_val(d, geo=g, time=y)
                if v is not None:
                    vv[y] = round(float(v), 1)
            if not vv:
                continue
            entry = {"code": g, "name": GEO_FR.get(g, gl.get(g, g)), "vals": vv}
            if g in ("EU27_2020", "EA"):
                aggs[g] = entry
            else:
                adoption.append(entry)
        adoption.sort(key=lambda e: -(e["vals"].get(ly) or 0))
        if adoption:
            out["adoption"] = adoption; out["aggregates"] = aggs
            got += 1; ok.append("eurostat:adoption")
        else:
            raise RuntimeError("aucun pays")
    except Exception as e:
        sys.stderr.write(f"[eurostat adoption] {e}\n"); failed.append("eurostat:adoption")
    # B) IA générative (image/vidéo/audio, E_AI_TPVSG) : geo × dernière année
    try:
        ly = out["latest_year"] or "2025"
        d = get_json(EUROSTAT_BASE + f"&indic_is=E_AI_TPVSG&time={ly}", timeout=40)
        dims = d["dimension"]; gl = dims["geo"]["category"]["label"]
        gidx = dims["geo"]["category"]["index"]; gen = []
        for g in gidx:
            v = _eurostat_val(d, geo=g)
            if v is None:
                continue
            gen.append({"code": g, "name": GEO_FR.get(g, gl.get(g, g)),
                        "pct": round(float(v), 1), "agg": g in ("EU27_2020", "EA")})
        gen.sort(key=lambda e: -e["pct"])
        if gen:
            out["generative"] = gen; got += 1; ok.append("eurostat:generative")
    except Exception as e:
        sys.stderr.write(f"[eurostat generative] {e}\n"); failed.append("eurostat:generative")
    # C) freins au non-usage + finalités d'usage : UE-27, dernière année
    try:
        ly = out["latest_year"] or "2025"
        d = get_json(EUROSTAT_BASE + f"&geo=EU27_2020&time={ly}", timeout=40)
        bar = []
        for code, label in BARRIERS_FR.items():
            v = _eurostat_val(d, indic_is=code)
            if v is not None:
                bar.append({"code": code, "label": label, "pct": round(float(v), 1)})
        bar.sort(key=lambda x: -x["pct"])
        pur = []
        for code, label in PURPOSES_FR.items():
            v = _eurostat_val(d, indic_is=code)
            if v is not None:
                pur.append({"code": code, "label": label, "pct": round(float(v), 1)})
        pur.sort(key=lambda x: -x["pct"])
        if bar:
            out["barriers_eu"] = bar
        if pur:
            out["purposes_eu"] = pur
        if bar or pur:
            got += 1; ok.append("eurostat:barriers_purposes")
    except Exception as e:
        sys.stderr.write(f"[eurostat barriers] {e}\n"); failed.append("eurostat:barriers_purposes")
    return out if got else None


# ════════════════════════════════════════════════════════════════════
#  Assemblage
# ════════════════════════════════════════════════════════════════════

def load_prev():
    try:
        return json.loads(OUT_JSON.read_text())
    except Exception:
        return {}

def build_payload():
    ok, failed = [], []
    prev = load_prev()

    # On part des valeurs de repli vérifiées, puis on écrase avec le live.
    agentic = json.loads(json.dumps(FALLBACK["agentic"]))
    intensity = json.loads(json.dumps(FALLBACK["intensity"]))
    enterprise = json.loads(json.dumps(FALLBACK["enterprise"]))
    public_admin = json.loads(json.dumps(FALLBACK["public_admin"]))
    personal = json.loads(json.dumps(FALLBACK["personal"]))
    labor = json.loads(json.dumps(FALLBACK["labor"]))
    geo = json.loads(json.dumps(FALLBACK["geo"]))

    # NB : on fusionne par clé (update) au lieu de remplacer, pour qu'un échec
    # partiel (un seul outil) conserve la valeur de repli des autres.
    npm = fetch_npm(ok, failed)
    if npm:
        agentic["npm"].update(npm)
    pypi = fetch_pypi(ok, failed)
    if pypi:
        agentic["pypi"].update(pypi)
    gh = fetch_github(ok, failed)
    if gh:
        agentic["github_stars"].update(gh)
        # historique stars : append diff quotidien
        hist = (prev.get("agentic", {}) or {}).get("stars_history", []) or []
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        snap = {"date": today, "stars": {r: v["stars"] for r, v in agentic["github_stars"].items()}}
        if not hist or hist[-1].get("date") != today:
            hist.append(snap)
        else:
            hist[-1] = snap
        agentic["stars_history"] = hist[-180:]

    orr = fetch_openrouter(ok, failed)
    if orr:
        intensity.update(orr)

    btos = fetch_btos(ok, failed)
    if btos:
        enterprise["btos"] = btos

    fed = fetch_federal_inventory(ok, failed)
    if fed:
        public_admin["us_federal_inventory"].update(fed)

    atrs = fetch_uk_atrs(ok, failed)
    if atrs:
        # historique compte ATRS
        hist = (prev.get("public_admin", {}) or {}).get("uk_atrs", {}).get("history", []) or []
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        if not hist or hist[-1].get("date") != today:
            hist.append({"date": today, "total": atrs["total"]})
        else:
            hist[-1] = {"date": today, "total": atrs["total"]}
        public_admin["uk_atrs"] = {"total": atrs["total"], "history": hist[-365:]}

    usa = fetch_usaspending(ok, failed)
    if usa:
        public_admin["usaspending_ai"] = usa

    ind = fetch_indeed(ok, failed)
    if ind:
        labor["indeed_ai_share"] = ind

    euro = fetch_eurostat_geo(ok, failed)
    if euro and euro.get("adoption"):
        geo["eurostat"] = euro

    meta = {
        "updated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "updated_at_unix": int(time.time()),
        "sources_ok": ok,
        "sources_failed": failed,
        "doc_version": "1.1",
        "note": "Toutes les séries live sont auditables : ouvrir les endpoints listés dans le bloc « Sources auditables » de l'onglet Adoption. Comparaison entre pays via Eurostat isoc_eb_ai.",
    }
    payload = {
        "meta": meta,
        "agentic": agentic,
        "intensity": intensity,
        "enterprise": enterprise,
        "public_admin": public_admin,
        "personal": personal,
        "labor": labor,
        "geo": geo,
    }
    return payload, len(ok), len(failed)


def write_outputs(payload):
    OUT_JSON.write_text(json.dumps(payload, separators=(",", ":"), ensure_ascii=False))
    js = (
        f"/* ai_adoption_cache.js — generated {payload['meta']['updated_at']} */\n"
        f"window.__AI_ADOPTION__ = "
        f"{json.dumps(payload, separators=(',', ':'), ensure_ascii=False)};\n"
    )
    OUT_JS.write_text(js)
    site_dir = Path.home() / "Desktop" / "Site_Crypto_Finance"
    if site_dir.exists():
        for name in ("ai_adoption_cache.json", "ai_adoption_cache.js"):
            link = site_dir / name
            target = CACHE_DIR / name
            try:
                if link.is_symlink() or link.exists():
                    link.unlink()
                link.symlink_to(target)
            except OSError as e:
                sys.stderr.write(f"[SYMLINK {name}] {e}, copie de secours\n")
                shutil.copy2(target, link)


def main():
    t0 = time.time()
    try:
        payload, n_ok, n_fail = build_payload()
    except Exception as e:
        sys.stderr.write(f"[FATAL] {e}\n")
        sys.exit(2)
    # garde-fou : si quasiment tout échoue et qu'un cache existe, on le garde
    if n_ok < 2 and OUT_JSON.exists():
        sys.stderr.write(f"[GUARD] {n_ok} OK / {n_fail} fail — conserve cache précédent\n")
        sys.exit(1)
    write_outputs(payload)
    dt = time.time() - t0
    sys.stdout.write(f"[ai_adoption] OK · {n_ok} sources OK, {n_fail} échec · {dt:.1f}s · {OUT_JSON}\n")


if __name__ == "__main__":
    main()
